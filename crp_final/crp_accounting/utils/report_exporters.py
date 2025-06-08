# crp_accounting/utils/report_exporters.py

import io
import logging
from datetime import date
from decimal import Decimal
from typing import List, Dict, Any, Optional, Tuple

# --- Excel ---
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# --- PDF ---
from django.template.loader import render_to_string # To render HTML templates
from xhtml2pdf import pisa                  # To generate PDF from HTML
# from django.conf import settings # Optional: To locate templates

# --- Graphs ---
import plotly.graph_objects as go

from ..models.coa import PLSection
# import plotly.io as pio # Not strictly needed if only writing bytes

# --- Local Type Imports (Adapt if your types are elsewhere) ---
# Assuming these are defined in services or a types file
from ..services.reports_service import (
    ProfitLossLineItem, ProfitLossAccountDetail,
    BalanceSheetNode
)

logger = logging.getLogger(__name__)

# =============================================================================
# General Styling Constants (Excel)
# =============================================================================
HEADER_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=12)
BOLD_FONT = Font(bold=True)
ACCOUNTING_FORMAT = '#,##0.00_);(#,##0.00)'
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical='top')
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')
RIGHT_ALIGNMENT = Alignment(horizontal='right', vertical='center')
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# =============================================================================
# Excel Generation Helpers
# =============================================================================

def _format_excel_header(sheet, report_title: str, report_info: Dict[str, Any]):
    """Adds standard header rows to an Excel sheet."""
    sheet.cell(row=1, column=1, value=report_title).font = HEADER_FONT
    row = 2
    for key, value in report_info.items():
        sheet.cell(row=row, column=1, value=f"{key}:").font = BOLD_FONT
        sheet.cell(row=row, column=2, value=value)
        row += 1
    # Add currency warning if applicable
    if 'report_currency' in report_info:
        sheet.cell(row=row-1, column=3, value="(Totals may mix currencies)").font = Font(italic=True, size=9)

    return row + 1 # Return next available row

def _add_excel_hierarchy_rows(
    sheet,
    nodes: List[Dict], # Expects BalanceSheetNode structure or similar TB structure
    start_row: int,
    balance_col_idx: int = 2, # Column index for the primary balance figure
    currency_col_idx: int = 3,
    indent_level: int = 0,
    level_indent_size: int = 2 # Spaces per indent level
) -> int:
    """Recursively adds hierarchical data to Excel rows."""
    current_row = start_row
    base_indent = " " * (indent_level * level_indent_size)

    for node in nodes:
        is_group = node.get('type') == 'group'
        indent = " " * (node.get('level', indent_level) * level_indent_size) # Use node's level if present

        # Column 1: Name (indented)
        name_cell = sheet.cell(row=current_row, column=1, value=f"{indent}{node['name']}")
        name_cell.alignment = Alignment(wrap_text=False, indent=node.get('level', indent_level)) # Use openpyxl indent
        if is_group: name_cell.font = BOLD_FONT

        # Column for Balance
        balance = node.get('balance', node.get('amount')) # Handle BS 'balance' or P&L 'amount'
        if balance is not None:
            balance_cell = sheet.cell(row=current_row, column=balance_col_idx, value=balance)
            balance_cell.number_format = ACCOUNTING_FORMAT
            balance_cell.alignment = RIGHT_ALIGNMENT
            if is_group: balance_cell.font = BOLD_FONT

        # Column for Currency (optional)
        if currency_col_idx and not is_group and node.get('currency'):
             sheet.cell(row=current_row, column=currency_col_idx, value=node['currency']).alignment = CENTER_ALIGNMENT

        current_row += 1
        # Recursively add children
        if node.get('children'):
            current_row = _add_excel_hierarchy_rows(
                sheet, node['children'], current_row,
                balance_col_idx, currency_col_idx, indent_level + 1, level_indent_size
            )
        # Recursively add accounts (for P&L structure)
        if node.get('accounts'):
             # Add a small indent for accounts under a P&L line
             account_indent = " " * ((indent_level + 1) * level_indent_size)
             for acc in node['accounts']:
                  acc_name_cell = sheet.cell(row=current_row, column=1, value=f"{account_indent}{acc['account_number']} - {acc['account_name']}")
                  acc_name_cell.alignment = Alignment(wrap_text=False, indent=indent_level+1)

                  acc_amount_cell = sheet.cell(row=current_row, column=balance_col_idx, value=acc['amount'])
                  acc_amount_cell.number_format = ACCOUNTING_FORMAT
                  acc_amount_cell.alignment = RIGHT_ALIGNMENT

                  if currency_col_idx and acc.get('currency'):
                       sheet.cell(row=current_row, column=currency_col_idx, value=acc['currency']).alignment = CENTER_ALIGNMENT
                  current_row += 1

    return current_row

def _auto_adjust_excel_columns(sheet, min_width=10, max_width=60):
    """Adjusts column widths based on content."""
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        adjusted_width = max(min_width, min(adjusted_width, max_width)) # Clamp width
        sheet.column_dimensions[column].width = adjusted_width

# =============================================================================
# Balance Sheet Exporters
# =============================================================================

def generate_balance_sheet_excel(report_data: Dict[str, Any]) -> bytes:
    """Generates a Balance Sheet report as an Excel file (.xlsx) in memory."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Balance Sheet {report_data.get('as_of_date')}"

    report_info = {
        "As of Date": report_data.get('as_of_date'),
        "Currency Context": report_data.get('report_currency'),
    }
    current_row = _format_excel_header(sheet, "Balance Sheet", report_info)

    # Define column indices
    COL_NAME = 1
    COL_BALANCE = 2
    COL_CURRENCY = 3

    # Assets
    sheet.cell(row=current_row, column=COL_NAME, value="ASSETS").font = SECTION_FONT
    current_row += 1
    assets_data = report_data.get('assets', {})
    current_row = _add_excel_hierarchy_rows(sheet, assets_data.get('hierarchy', []), current_row, COL_BALANCE, COL_CURRENCY)
    sheet.cell(row=current_row, column=COL_NAME, value="Total Assets").font = BOLD_FONT
    cell = sheet.cell(row=current_row, column=COL_BALANCE, value=assets_data.get('total', 0))
    cell.font = BOLD_FONT; cell.number_format = ACCOUNTING_FORMAT; cell.alignment = RIGHT_ALIGNMENT
    current_row += 2

    # Liabilities
    sheet.cell(row=current_row, column=COL_NAME, value="LIABILITIES").font = SECTION_FONT
    current_row += 1
    liabilities_data = report_data.get('liabilities', {})
    current_row = _add_excel_hierarchy_rows(sheet, liabilities_data.get('hierarchy', []), current_row, COL_BALANCE, COL_CURRENCY)
    sheet.cell(row=current_row, column=COL_NAME, value="Total Liabilities").font = BOLD_FONT
    cell = sheet.cell(row=current_row, column=COL_BALANCE, value=liabilities_data.get('total', 0))
    cell.font = BOLD_FONT; cell.number_format = ACCOUNTING_FORMAT; cell.alignment = RIGHT_ALIGNMENT
    current_row += 2

    # Equity
    sheet.cell(row=current_row, column=COL_NAME, value="EQUITY").font = SECTION_FONT
    current_row += 1
    equity_data = report_data.get('equity', {})
    current_row = _add_excel_hierarchy_rows(sheet, equity_data.get('hierarchy', []), current_row, COL_BALANCE, COL_CURRENCY) # Includes RE
    sheet.cell(row=current_row, column=COL_NAME, value="Total Equity").font = BOLD_FONT
    cell = sheet.cell(row=current_row, column=COL_BALANCE, value=equity_data.get('total', 0))
    cell.font = BOLD_FONT; cell.number_format = ACCOUNTING_FORMAT; cell.alignment = RIGHT_ALIGNMENT
    current_row += 2

    # Balance Check Summary
    sheet.cell(row=current_row, column=COL_NAME, value="Total Liabilities + Equity").font = BOLD_FONT
    cell = sheet.cell(row=current_row, column=COL_BALANCE, value=liabilities_data.get('total', 0) + equity_data.get('total', 0))
    cell.font = BOLD_FONT; cell.number_format = ACCOUNTING_FORMAT; cell.alignment = RIGHT_ALIGNMENT
    current_row += 1
    sheet.cell(row=current_row, column=COL_NAME, value="Balanced Check").font = BOLD_FONT
    sheet.cell(row=current_row, column=COL_BALANCE, value="Balanced" if report_data.get('is_balanced') else "OUT OF BALANCE").font = BOLD_FONT

    _auto_adjust_excel_columns(sheet, max_width=70) # Adjust after adding all data

    # Save to buffer
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_balance_sheet_pdf(report_data: Dict[str, Any]) -> bytes:
    """Generates a Balance Sheet report as a PDF file in memory using HTML templates."""
    # --- Template Approach ---
    template_path = 'reports/balance_sheet_pdf.html' # Example path relative to template dirs
    context = {'report': report_data} # Pass data to the template

    try:
        html = render_to_string(template_path, context)
        # You might need to define CSS separately or include it in the template
        # css_path = os.path.join(settings.STATICFILES_DIRS[0], 'css/report_style.css')
        # with open(css_path, 'r') as f:
        #     css_string = f.read()

        buffer = io.BytesIO()
        # Generate PDF
        pisa_status = pisa.CreatePDF(
            src=html,                # Source HTML
            dest=buffer,             # File handle to recieve result
            # link_callback=link_callback # Optional handler for images/static files if needed
            # css=css_string           # Optional CSS string
        )

        if pisa_status.err:
            logger.error(f"PDF generation error for Balance Sheet {report_data.get('as_of_date')}: {pisa_status.err}")
            raise RuntimeError(f"PDF generation failed: {pisa_status.err}")

        buffer.seek(0)
        return buffer.getvalue()

    except Exception as e:
        logger.exception(f"Error rendering or generating Balance Sheet PDF: {e}")
        raise # Re-raise the exception


# =============================================================================
# Profit & Loss Exporters
# =============================================================================

def generate_profit_loss_excel(report_data: Dict[str, Any]) -> bytes:
    """Generates a Profit & Loss report as an Excel file (.xlsx) in memory."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Profit & Loss {report_data.get('start_date')} to {report_data.get('end_date')}"

    report_info = {
        "Start Date": report_data.get('start_date'),
        "End Date": report_data.get('end_date'),
        "Currency Context": report_data.get('report_currency'),
    }
    current_row = _format_excel_header(sheet, "Profit and Loss Statement", report_info)

    # Define column indices
    COL_NAME = 1
    COL_AMOUNT = 2
    COL_CURRENCY = 3

    # Add P&L lines (handling nesting of accounts under sections)
    current_row = _add_excel_hierarchy_rows(
        sheet,
        report_data.get('report_lines', []),
        current_row,
        balance_col_idx=COL_AMOUNT, # P&L uses 'amount' key
        currency_col_idx=COL_CURRENCY,
        indent_level=0,
        level_indent_size=2 # Smaller indent maybe
    )

    # Add final Net Income line explicitly if needed, or ensure it's formatted in loop
    # The loop above should handle it if NET_INCOME is the last item

    _auto_adjust_excel_columns(sheet, max_width=70)

    # Save to buffer
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_profit_loss_pdf(report_data: Dict[str, Any]) -> bytes:
    """Generates a Profit & Loss report as a PDF file in memory using HTML templates."""
    template_path = 'reports/profit_loss_pdf.html' # Example path
    context = {'report': report_data}
    try:
        html = render_to_string(template_path, context)
        buffer = io.BytesIO()
        pisa_status = pisa.CreatePDF(src=html, dest=buffer)
        if pisa_status.err:
            logger.error(f"PDF generation error for P&L {report_data.get('start_date')}-{report_data.get('end_date')}: {pisa_status.err}")
            raise RuntimeError(f"PDF generation failed: {pisa_status.err}")
        buffer.seek(0)
        return buffer.getvalue()
    except Exception as e:
        logger.exception(f"Error rendering or generating P&L PDF: {e}")
        raise


# =============================================================================
# Graph Generation (Plotly - Returning Image Bytes)
# =============================================================================

def _create_pl_waterfall_figure(report_lines: List[ProfitLossLineItem]) -> Optional[go.Figure]:
    """Creates a Plotly Waterfall chart figure for P&L."""
    measures = []
    values = []
    texts = []
    y_labels = []

    # Extract relevant data - adjust based on how detailed you want it
    # This example uses major sections/subtotals
    keys_to_include = [
        PLSection.REVENUE.value, PLSection.COGS.value, 'GROSS_PROFIT',
        PLSection.OPERATING_EXPENSE.value, 'OPERATING_PROFIT', # Add others as needed
        'NET_INCOME'
    ]
    section_map = {line['section_key']: line for line in report_lines}
    last_value = 0

    for key in keys_to_include:
        if key in section_map:
            line = section_map[key]
            value = line['amount']
            y_labels.append(line['title'])

            if line['is_subtotal']:
                measures.append("total")
                values.append(value)
                texts.append(f"{value:,.2f}")
                last_value = value
            else:
                # Calculate change from last subtotal/value
                delta = value
                # Adjust delta sign based on impact (Expenses/COGS reduce profit)
                if key in [PLSection.COGS.value, PLSection.OPERATING_EXPENSE.value, PLSection.TAX_EXPENSE.value, PLSection.OTHER_EXPENSE.value]:
                    delta = -value # Waterfall expects negative for decreases

                measures.append("relative")
                values.append(delta)
                texts.append(f"{delta:+,.2f}") # Show sign for relative changes
                last_value += delta # This is conceptually wrong for waterfall value calc, need to fix waterfall logic if used.
                                    # Waterfall values usually represent the change. Totals are automatic.

    # Basic Waterfall - Needs refinement for accurate value progression
    if not y_labels: return None
    fig = go.Figure(go.Waterfall(
        name = "P&L", orientation = "v",
        measure = measures, # ["relative", "relative", "total", "relative", "total", ...]
        x = y_labels,
        textposition = "outside",
        # text = texts, # Let Plotly calculate text from y
        y = values, # Should contain the delta values for relative, absolute for total
        connector = {"line":{"color":"rgb(63, 63, 63)"}},
    ))
    fig.update_layout(title="Profit & Loss Waterfall", showlegend=False)
    return fig


def generate_profit_loss_waterfall_png(report_data: Dict[str, Any]) -> Optional[bytes]:
    """Generates a P&L Waterfall chart as PNG bytes."""
    fig = _create_pl_waterfall_figure(report_data.get('report_lines', []))
    if fig:
        try:
            buffer = io.BytesIO()
            fig.write_image(buffer, format="png", scale=2) # Use Kaleido
            buffer.seek(0)
            return buffer.getvalue()
        except Exception as e:
            logger.error(f"Failed to generate P&L waterfall image: {e}")
            # This might happen if Kaleido is not installed or configured correctly
    return None

# --- Add more graph functions as needed (e.g., BS composition pie charts) ---
# def generate_asset_composition_pie_png(report_data: ...) -> Optional[bytes]:
#     pass


# =============================================================================
# Trial Balance Exporters (Placeholder - Adapt as needed)
# =============================================================================

def generate_trial_balance_excel(report_data: Dict[str, Any]) -> bytes:
    """Generates a Trial Balance report as an Excel file (.xlsx) in memory."""
    # --- Implementation similar to Balance Sheet but with Dr/Cr columns ---
    # You would adapt the _add_excel_hierarchy_rows or write a specific one
    # Use the 'hierarchy' data from generate_trial_balance_structured
    # Columns: Account Number, Account Name, Debit, Credit
    workbook = Workbook()
    sheet = workbook.active
    # ... add headers, iterate hierarchy, format ...
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    logger.warning("Trial Balance Excel export is a placeholder implementation.")
    return buffer.getvalue() # Return empty/basic file for now

def generate_trial_balance_pdf(report_data: Dict[str, Any]) -> bytes:
    """Generates a Trial Balance report as a PDF file in memory using HTML templates."""
    # --- Implementation similar to P&L/BS PDF ---
    template_path = 'reports/trial_balance_pdf.html' # Example path
    # ... render template, generate PDF ...
    buffer = io.BytesIO()
    logger.warning("Trial Balance PDF export is a placeholder implementation.")
    # pisa.CreatePDF(...)
    buffer.seek(0)
    return buffer.getvalue() # Return empty PDF for now

# =============================================================================
# --- End of File ---
# =============================================================================