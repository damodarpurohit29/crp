# crp_accounting/admin/period.py

import logging
from io import BytesIO  # For Excel export, if kept
from urllib.parse import urlencode  # For report links
from decimal import Decimal
from typing import Optional, Any, Tuple  # For type hinting

from django.contrib import admin, messages
from django.contrib.admin.widgets import AdminDateWidget
from django.db import models as django_db_models, models
from django.urls import reverse, NoReverseMatch
from django.utils.html import format_html
from django.utils.translation import gettext_lazy as _
from django.http import HttpResponse, HttpRequest
from django.core.exceptions import ValidationError as DjangoValidationError
from django.utils.http import unquote  # For object_id from URL

from ..exceptions import ReportGenerationError
# --- App-Specific Imports ---
from ..models.period import FiscalYear, AccountingPeriod
from ..serializers import FiscalYearStatus
from ..services import reports_service  # If used by actions (like export_trial_balance)
from .admin_base import TenantAccountingModelAdmin  # Your base tenant-aware admin class
from company.models import Company  # For type hinting and direct use

# --- Excel Library (Conditional Import) ---
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("Admin Period: 'openpyxl' library not found. Excel exports will be disabled.")

logger = logging.getLogger("crp_accounting.admin.period")  # Specific logger for this admin file
ZERO = Decimal('0.00')  # If used in actions, else can be removed


# =============================================================================
# Fiscal Year Admin
# =============================================================================
@admin.register(FiscalYear)
class FiscalYearAdmin(TenantAccountingModelAdmin):
    # list_display: 'company' column is added dynamically by TenantAccountingModelAdmin for superusers
    list_display = ('name', 'start_date', 'end_date', 'status_colored', 'is_active_display', 'closed_by_user_display',
                    'updated_at_short')

    list_filter_non_superuser = ('status', 'is_active', ('start_date', admin.DateFieldListFilter))
    # Superuser filter will be constructed in get_list_filter

    search_fields = ('name', 'company__name')  # Allow superuser to search by company name

    # These fields are typically results of actions or system-set, not directly editable on change.
    readonly_fields_on_change = ('status', 'closed_by', 'closed_at')
    # Audit fields like created_at, created_by, updated_at, updated_by are handled by TenantAccountingModelAdmin.get_readonly_fields

    list_select_related = ('company', 'closed_by')  # For efficient list display
    actions = ["admin_action_activate_year", "admin_action_lock_year", "admin_action_close_year",
               "admin_action_reopen_year"]
    ordering = ('-company__name', '-start_date',)  # For SUs: group by company, then newest FY

    formfield_overrides = {
        django_db_models.DateField: {'widget': AdminDateWidget},  # Use Django's nicer date picker
    }

    def get_list_filter(self, request: HttpRequest) -> tuple:
        """Dynamically sets list_filter based on whether the user is a superuser."""
        if request.user.is_superuser:
            return ('company',) + self.list_filter_non_superuser  # Superuser gets 'company' filter
        return self.list_filter_non_superuser

    def get_fieldsets(self, request: HttpRequest, obj: Optional[FiscalYear] = None) -> tuple:
        """
        Defines the layout of fields on the add/change form.
        The 'company' field is always included in the structure; its editability and initial value
        are controlled by TenantAccountingModelAdmin's get_readonly_fields and save_model/initial_data.
        """
        detail_fields = ('name', 'start_date', 'end_date')

        # 'company' field is always part of the first fieldset.
        # - For non-SU on add: readonly and pre-filled.
        # - For SU on add: editable dropdown.
        # - For all users on change: readonly.
        main_fieldset_fields = ('company',) + detail_fields

        if obj is None:  # Add view
            return (
                (_('Fiscal Year Details'), {'fields': main_fieldset_fields}),
                # Status and audit info are not typically set on initial creation directly by user.
            )
        else:  # Change view
            status_fields = ('status', 'is_active', 'closed_by', 'closed_at')
            # Assuming 'created_by' and 'updated_by' exist on TenantScopedModel or FiscalYear
            audit_fields = ('created_at', 'updated_at', 'created_by', 'updated_by')
            return (
                (_('Fiscal Year Details'), {'fields': main_fieldset_fields}),
                (_('Status Information'), {'fields': status_fields}),
                (_('Audit Information'), {'fields': audit_fields, 'classes': ('collapse',)}),
            )

    def get_readonly_fields(self, request: HttpRequest, obj: Optional[FiscalYear] = None) -> tuple:
        """Determines which fields are read-only on the form."""
        # Start with readonly fields from TenantAccountingModelAdmin (handles 'company' correctly)
        ro_fields = set(super().get_readonly_fields(request, obj))
        if obj:  # If it's a change view (editing an existing FiscalYear)
            ro_fields.update(self.readonly_fields_on_change)
        return tuple(ro_fields)

    # --- Display methods for list_display ---
    @admin.display(description=_('Status'), ordering='status')
    def status_colored(self, obj: FiscalYear):  # Renamed for clarity
        color_map = {
            FiscalYearStatus.OPEN: "green",
            FiscalYearStatus.LOCKED: "orange",
            FiscalYearStatus.CLOSED: "red",
        }
        # Ensure obj.get_status_display() is available (Django default for choice fields)
        return format_html(
            f'<span style="color:{color_map.get(obj.status, "grey")}; font-weight:bold;">{obj.get_status_display()}</span>')

    @admin.display(description=_('Active'), boolean=True, ordering='is_active')
    def is_active_display(self, obj: FiscalYear):
        return obj.is_active  # Assumes FiscalYear has an 'is_active' boolean field

    @admin.display(description=_('Closed By'), ordering='closed_by__name')
    def closed_by_user_display(self, obj: FiscalYear):
        return obj.closed_by.name if obj.closed_by_id else "N/A"  # Access name safely

    @admin.display(description=_('Last Updated'), ordering='updated_at')
    def updated_at_short(self, obj: FiscalYear):
        return obj.updated_at.strftime('%Y-%m-%d %H:%M') if obj.updated_at else "N/A"

    # --- Admin Actions ---
    # (Your existing admin action methods: admin_action_activate_year,
    #  admin_action_lock_year, admin_action_close_year, admin_action_reopen_year.
    #  Ensure they correctly call model methods which should handle their own logic,
    #  including company context if necessary, though usually they operate on `self.company`.)
    @admin.action(description=_('Activate selected Fiscal Year (deactivates others in same company)'))
    def admin_action_activate_year(self, request: HttpRequest, queryset: models.QuerySet):
        if queryset.count() != 1: self.message_user(request, _("Please select exactly one Fiscal Year to activate."),
                                                    messages.WARNING); return
        year = queryset.first()
        try:
            year.activate(); self.message_user(request,
                                               _("Fiscal Year '%(name)s' for Company '%(company)s' activated successfully.") % {
                                                   'name': year.name, 'company': year.company.name}, messages.SUCCESS)
        except DjangoValidationError as e:
            self.message_user(request, f"Error activating '{year.name}': {e.messages_joined}", messages.ERROR)
        except Exception as e:
            logger.error(f"Admin error activating FY {year.pk} (Co: {year.company_id}): {e}",
                         exc_info=True); self.message_user(request, _("An unexpected error occurred: %(error)s") % {
                'error': str(e)}, messages.ERROR)

    @admin.action(description=_('Lock selected OPEN Fiscal Years'))
    def admin_action_lock_year(self, request: HttpRequest, queryset: models.QuerySet):
        updated_count = 0
        for year in queryset.filter(status=FiscalYearStatus.OPEN.value):
            try:
                year.lock_year(); updated_count += 1
            except DjangoValidationError as e:
                self.message_user(request, _("Could not lock Fiscal Year '%(name)s': %(error)s") % {'name': year.name,
                                                                                                    'error': e.messages_joined},
                                  messages.ERROR)
        if updated_count > 0: self.message_user(request,
                                                _("%(count)d fiscal year(s) locked.") % {'count': updated_count},
                                                messages.SUCCESS)

    @admin.action(description=_('Close selected LOCKED Fiscal Years'))
    def admin_action_close_year(self, request: HttpRequest, queryset: models.QuerySet):
        updated_count = 0
        for year in queryset.filter(status=FiscalYearStatus.LOCKED.value):
            try:
                year.close_year(user=request.user); updated_count += 1  # Pass user if model method expects it
            except DjangoValidationError as e:
                self.message_user(request, _("Could not close Fiscal Year '%(name)s': %(error)s") % {'name': year.name,
                                                                                                     'error': e.messages_joined},
                                  messages.ERROR)
        if updated_count > 0: self.message_user(request,
                                                _("%(count)d fiscal year(s) closed.") % {'count': updated_count},
                                                messages.SUCCESS)

    @admin.action(description=_('Reopen selected Fiscal Years (LOCKED/CLOSED to OPEN) - Use Caution!'))
    def admin_action_reopen_year(self, request: HttpRequest, queryset: models.QuerySet):
        updated_count = 0
        for year in queryset.exclude(status=FiscalYearStatus.OPEN.value):
            try:
                year.reopen_year(); updated_count += 1
            except DjangoValidationError as e:
                self.message_user(request, _("Could not reopen Fiscal Year '%(name)s': %(error)s") % {'name': year.name,
                                                                                                      'error': e.messages_joined},
                                  messages.ERROR)
        if updated_count > 0: self.message_user(request, _("%(count)d fiscal year(s) reopened to OPEN status.") % {
            'count': updated_count}, messages.WARNING)


# =============================================================================
# Accounting Period Admin
# =============================================================================
@admin.register(AccountingPeriod)
class AccountingPeriodAdmin(TenantAccountingModelAdmin):
    list_display = (
    'name', 'fiscal_year_display_list', 'start_date', 'end_date', 'lock_status_display', 'view_reports_links',
    'updated_at_short')
    list_filter_non_superuser = (
    'locked', ('fiscal_year', admin.RelatedOnlyFieldListFilter), ('start_date', admin.DateFieldListFilter))
    # Superuser filter constructed in get_list_filter

    search_fields = ('name', 'fiscal_year__name', 'company__name')
    readonly_fields_on_change = ('locked',)  # Basic readonly for change view
    list_select_related = ('company', 'fiscal_year__company')  # Crucial for performance and display methods
    autocomplete_fields = ['company', 'fiscal_year']  # 'company' is for SU add view
    actions = ["admin_action_lock_periods", "admin_action_unlock_periods", "admin_action_export_trial_balance"]
    ordering = ('-fiscal_year__company__name', '-fiscal_year__start_date', '-start_date',)  # Ensures logical ordering

    formfield_overrides = {
        django_db_models.DateField: {'widget': AdminDateWidget},
    }

    def get_list_filter(self, request: HttpRequest) -> tuple:
        """Dynamically sets list_filter based on user type."""
        if request.user.is_superuser:
            # SU can filter by the Period's direct company, and also by the Fiscal Year's company (should be same).
            return (
            'company', ('fiscal_year__company', admin.RelatedOnlyFieldListFilter)) + self.list_filter_non_superuser
        return self.list_filter_non_superuser

    def get_fieldsets(self, request: HttpRequest, obj: Optional[AccountingPeriod] = None) -> tuple:
        """Defines form layout, ensuring 'company' field is structurally present."""
        detail_fields = ('name', 'fiscal_year', 'start_date', 'end_date')
        main_fieldset_fields = ('company',) + detail_fields  # 'company' always included here

        if obj is None:  # Add view
            return (
                (_('Period Details'), {'fields': main_fieldset_fields}),
            )
        else:  # Change view
            status_fields = ('locked',)
            audit_fields = ('created_at', 'updated_at', 'created_by', 'updated_by')
            return (
                (_('Period Details'), {'fields': main_fieldset_fields}),
                (_('Status Information'), {'fields': status_fields}),
                (_('Audit Information'), {'fields': audit_fields, 'classes': ('collapse',)}),
            )

    def get_readonly_fields(self, request: HttpRequest, obj: Optional[AccountingPeriod] = None) -> tuple:
        """Determines read-only fields, inheriting from base and adding specifics."""
        ro_fields = set(super().get_readonly_fields(request, obj))
        if obj:  # Change view
            ro_fields.update(self.readonly_fields_on_change)
            ro_fields.add('fiscal_year')  # Fiscal year typically not changed after period creation
        return tuple(ro_fields)

    def formfield_for_foreignkey(self, db_field: models.ForeignKey, request: HttpRequest, **kwargs: Any) -> Optional[
        models.Field]:
        """
        Filters ForeignKey choices, especially for 'fiscal_year'.
        Crucially ensures 'fiscal_year' choices are filtered by the company of the
        AccountingPeriod being added/edited AND pre-fetches fiscal_year.company.
        """
        # Determine the company context FOR THE AccountingPeriod OBJECT being manipulated.
        current_period_company_context: Optional[Company] = self._get_company_from_request_obj_or_form(
            request,
            obj=self.get_object(request, unquote(
                request.resolver_match.kwargs.get('object_id', ''))) if request.resolver_match.kwargs.get(
                'object_id') else None,
            form_data_for_add_view_post=request.POST if request.method == 'POST' and not request.resolver_match.kwargs.get(
                'object_id') else None
        )
        log_prefix_ffk = f"[APeriodAdmin FFKey][User:{request.user.name}][Fld:'{db_field.name}']"
        logger.debug(
            f"{log_prefix_ffk} Period's CoCtx: {current_period_company_context.name if current_period_company_context else 'None'}")

        if db_field.name == "fiscal_year":
            if current_period_company_context:
                # Filter FiscalYear choices to those belonging to the current_period_company_context.
                # Also, select_related('company') on FiscalYear is VITAL here to prevent
                # RelatedObjectDoesNotExist when FiscalYear.__str__ (which might access fiscal_year.company.name) is called.
                kwargs["queryset"] = FiscalYear.objects.filter(
                    company=current_period_company_context
                    # status=FiscalYearStatus.OPEN.value # Optionally filter for OPEN Fiscal Years only
                ).select_related('company').order_by('-start_date')
                logger.debug(
                    f"{log_prefix_ffk} Filtered FiscalYear choices for Co '{current_period_company_context.name}'. Count: {kwargs['queryset'].count()}.")
            else:
                # No company context determined for the AccountingPeriod (e.g., SU adding, 'company' field not yet selected).
                kwargs["queryset"] = FiscalYear.objects.none()
                logger.warning(f"{log_prefix_ffk} No CoCtx for current Period, so FiscalYear queryset is None.")
                if not request.resolver_match.kwargs.get('object_id') and request.user.is_superuser:
                    messages.info(request, _("Select 'Company' (for this new Period) to populate Fiscal Year choices."))

        # Let TenantAccountingModelAdmin handle other fields or general company filtering if db_field.name == 'company'
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    # --- Display methods ---
    @admin.display(description=_('Fiscal Year'), ordering='fiscal_year__name')
    def fiscal_year_display_list(self, obj: AccountingPeriod):  # Renamed to avoid conflict if form display different
        if obj.fiscal_year_id:
            # obj.fiscal_year is available due to list_select_related=('fiscal_year__company')
            return str(obj.fiscal_year) if obj.fiscal_year else "N/A (FY Link Error)"
        return "N/A"

    @admin.display(description=_("Status"), ordering='locked')
    def lock_status_display(self, obj: AccountingPeriod):
        return format_html('<span style="color:red; font-weight:bold;">🔒 Locked</span>') if obj.locked \
            else format_html('<span style="color:green; font-weight:bold;">🔓 Open</span>')

    @admin.display(description=_('Last Updated'), ordering='updated_at')
    def updated_at_short(self, obj: AccountingPeriod):
        return obj.updated_at.strftime('%Y-%m-%d %H:%M') if obj.updated_at else "N/A"

    # --- Admin Actions for AccountingPeriod ---
    @admin.action(description=_('Lock selected OPEN periods'))
    def admin_action_lock_periods(self, request: HttpRequest, queryset: models.QuerySet):
        updated_count = 0
        for period in queryset.filter(locked=False):
            try:
                period.lock_period(); updated_count += 1
            except DjangoValidationError as e:
                self.message_user(request, _("Could not lock period '%(name)s': %(error)s") % {'name': period.name,
                                                                                               'error': e.messages_joined},
                                  messages.ERROR)
        if updated_count > 0: self.message_user(request, _("%(count)d period(s) locked.") % {'count': updated_count},
                                                messages.SUCCESS)

    @admin.action(description=_('Unlock selected LOCKED periods'))
    def admin_action_unlock_periods(self, request: HttpRequest, queryset: models.QuerySet):
        updated_count = 0
        for period in queryset.filter(locked=True):
            try:
                period.unlock_period(); updated_count += 1
            except DjangoValidationError as e:
                self.message_user(request, _("Could not unlock period '%(name)s': %(error)s") % {'name': period.name,
                                                                                                 'error': e.messages_joined},
                                  messages.ERROR)
        if updated_count > 0: self.message_user(request, _("%(count)d period(s) unlocked.") % {'count': updated_count},
                                                messages.SUCCESS)

    @admin.display(description=_('View Reports'))
    def view_reports_links(self, obj: AccountingPeriod):  # obj is an AccountingPeriod instance
        if not obj.company_id or not obj.fiscal_year_id:  # company_id comes from the AccountingPeriod model
            return "N/A (Missing Company/FY Info)"

        report_links = []
        namespace = "crp_accounting_api"  # This is your app_name from urls_api.py

        # --- MODIFICATION START ---
        # Add company_id to the base parameters for all report links
        # Assuming your AccountingPeriod model has a `company_id` field or `company.pk`
        # If AccountingPeriod has a direct ForeignKey to Company named 'company':
        # company_pk_to_use = obj.company.pk if obj.company else None
        # If AccountingPeriod just has 'company_id':
        company_pk_to_use = obj.company_id

        if not company_pk_to_use:
            logger.warning(f"AccountingPeriod {obj.pk} is missing company information for report links.")
            return "N/A (Period missing company link)"

        base_link_params = {'company_id': company_pk_to_use}

        # --- MODIFICATION END ---

        def build_url(view_name_suffix, params):
            try:
                # Construct the URL name, e.g., "crp_accounting_api:admin-view-balance-sheet"
                url_name = f"{namespace}:admin-view-{view_name_suffix.replace('_', '-')}"
                base_url = reverse(url_name)  # Reverse without args if your report URLs don't have path params

                # Append query parameters
                return f"{base_url}?{urlencode(params)}"
            except NoReverseMatch:
                logger.warning(
                    f"Admin Period: URL not found via reverse for '{url_name}'. Check urls.py and namespace.")
                return None

        # Trial Balance
        # Merge base_link_params with report-specific params
        tb_params = {**base_link_params, 'as_of_date': obj.end_date.isoformat()}
        tb_url = build_url('trial-balance', tb_params)
        if tb_url:
            report_links.append(format_html('<a href="{}" target="_blank">TB</a>', tb_url))

        # Profit & Loss
        pl_params = {
            **base_link_params,
            'start_date': obj.start_date.isoformat(),
            'end_date': obj.end_date.isoformat()
        }
        pl_url = build_url('profit-loss', pl_params)
        if pl_url:
            report_links.append(format_html('<a href="{}" target="_blank">P&L</a>', pl_url))

        # Balance Sheet
        bs_params = {**base_link_params, 'as_of_date': obj.end_date.isoformat()}
        bs_url = build_url('balance-sheet', bs_params)
        if bs_url:
            report_links.append(format_html('<a href="{}" target="_blank">BS</a>', bs_url))

        return format_html(' | '.join(report_links)) if report_links else "No Reports Configured"

    @admin.action(description=_('Export Trial Balance to Excel for selected period(s)'))
    def admin_action_export_trial_balance(self, request, queryset):
        if not OPENPYXL_AVAILABLE:
            self.message_user(request, _("Excel export requires 'openpyxl' library to be installed."), messages.ERROR)
            return
        if queryset.count() != 1:
            self.message_user(request, _("Please select exactly one accounting period for this export action."),
                              messages.WARNING)
            return
        period = queryset.select_related('company').first()
        if not period or not period.company_id:
            self.message_user(request, _("Cannot export: Valid Company association not found for the period."),
                              messages.ERROR)
            return

        company_id_for_report = period.company_id
        as_of_date_for_report = period.end_date
        company_name_for_filename = period.company.name if period.company else f"Company_{company_id_for_report}"
        try:
            report_data = reports_service.generate_trial_balance_structured(
                company_id=company_id_for_report, as_of_date=as_of_date_for_report
            )
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = f"TB {as_of_date_for_report.strftime('%Y-%m-%d')}"
            bold_font = Font(bold=True, size=12)
            sheet['A1'] = f"Trial Balance - {company_name_for_filename}";
            sheet['A1'].font = Font(bold=True, size=14)
            sheet.merge_cells('A1:D1')
            sheet['A2'] = f"As of: {as_of_date_for_report.strftime('%B %d, %Y')}"
            sheet['A3'] = f"Currency: {report_data.get('report_currency', 'N/A')}"
            headers = ["Acc. No.", "Account Name", "Debit", "Credit"]
            sheet.append([]);
            header_row_idx = sheet.max_row + 1
            for col_num, header_text in enumerate(headers, 1):
                cell = sheet.cell(row=header_row_idx, column=col_num, value=header_text);
                cell.font = bold_font
                sheet.column_dimensions[get_column_letter(col_num)].width = 18 if col_num in [3, 4] else (
                    40 if col_num == 2 else 15)
            for entry in report_data.get('flat_entries', []):
                sheet.append([entry.get('account_number', ''), entry.get('account_name', ''), entry.get('debit', ZERO),
                              entry.get('credit', ZERO)])
                for col_idx in [3, 4]: sheet.cell(row=sheet.max_row, column=col_idx).number_format = '#,##0.00'
            sheet.append([]);
            total_row_idx = sheet.max_row + 1
            sheet.cell(row=total_row_idx, column=2, value="TOTALS:").font = bold_font
            sheet.cell(row=total_row_idx, column=3, value=report_data.get('total_debit', ZERO)).font = bold_font
            sheet.cell(row=total_row_idx, column=3).number_format = '#,##0.00'
            sheet.cell(row=total_row_idx, column=4, value=report_data.get('total_credit', ZERO)).font = bold_font
            sheet.cell(row=total_row_idx, column=4).number_format = '#,##0.00'
            if not report_data.get('is_balanced', True):
                sheet.cell(row=total_row_idx + 1, column=1,
                           value="WARNING: Trial Balance is NOT balanced!").font = Font(color="FF0000", bold=True)
            response_stream = BytesIO()
            workbook.save(response_stream);
            response_stream.seek(0)
            response = HttpResponse(response_stream.read(),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            filename = f"{company_name_for_filename}_Trial_Balance_{period.name.replace(' ', '_')}_{as_of_date_for_report.strftime('%Y%m%d')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        except DjangoValidationError as ve:
            self.message_user(request, f"Export failed due to validation: {ve.messages_joined}", messages.ERROR)
        except Exception as e:
            logger.exception(f"Error exporting Trial Balance for Co {company_id_for_report}, Period {period.pk}: {e}")
            self.message_user(request, f"An unexpected error occurred during export: {str(e)}", messages.ERROR)