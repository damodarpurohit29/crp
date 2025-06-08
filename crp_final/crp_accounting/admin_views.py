import logging
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Union, Any, List, Dict, Optional
import os  # For link_callback

from django.conf import settings  # For link_callback
from django.contrib import messages
from django.shortcuts import render, get_object_or_404
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.admin import site as admin_site
from django.core.exceptions import ValidationError as DjangoValidationError, PermissionDenied as DjangoPermissionDenied
from django.http import HttpRequest, HttpResponse, Http404, HttpResponseBadRequest  # Ensure HttpRequest is imported
from django.template.loader import render_to_string
from django.utils.translation import gettext_lazy as _

# --- Model Imports ---
from .models.coa import Account
from .models.journal import VoucherType
from .models.party import Party
from .services.voucher_service import PERMISSIONS_MAP, get_permissions_for_role

# --- Company Model Import ---
try:
    from company.models import Company, CompanyMembership
except ImportError:
    Company = None
    logging.critical("Admin Views: CRITICAL - Could not import Company model. Tenant context will fail.")

# --- Enum Imports ---
from crp_core.enums import AccountNature, \
    PartyType as CorePartyType

# --- PDF/Excel Imports ---
try:
    from xhtml2pdf import pisa

    XHTML2PDF_AVAILABLE = True
except ImportError:
    XHTML2PDF_AVAILABLE = False
    logging.warning("Admin Views: xhtml2pdf library not found. PDF export will not be available.")
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("Admin Views: openpyxl library not found. Excel export will not be available.")

# --- Service Imports ---
from .services import reports_service, ledger_service
from .exceptions import ReportGenerationError

logger = logging.getLogger("crp_accounting.admin_views")
ZERO = Decimal('0.00')
PK_TYPE = Union[int, str, 'uuid.UUID']


# =========================================
# Helper Functions
# =========================================
def _get_admin_base_context(title: str, request: HttpRequest) -> Dict[str, Any]:
    resolver_match_opts = getattr(request.resolver_match, 'opts', None)
    current_app = getattr(request.resolver_match, 'app_name', admin_site.name)
    return {
        'title': title,
        'site_header': admin_site.site_header,
        'site_title': admin_site.site_title,
        'opts': resolver_match_opts or Account._meta,
        'has_permission': True,
        'is_popup': request.GET.get('_popup') == '1',
        'is_nav_sidebar_enabled': not (request.GET.get('_popup') == '1'),
        'current_app': current_app,
        'user': request.user,
        # 'request': request, # No need to add request here if render() or render_to_string(..., request=request) is used
    }


def _get_company_for_report_or_raise(request: HttpRequest) -> Company:
    log_prefix = f"[GetCoForReport][User:{getattr(request.user, 'name', request.user.id)}]"
    if not Company:
        logger.critical(f"{log_prefix} Company model not available.")
        raise Http404(_("System configuration error: Company module not available."))

    company_from_middleware = getattr(request, 'company', None)

    if request.user.is_superuser:
        company_id_str = request.GET.get('company_id')
        if company_id_str:
            try:
                company_pk = Company._meta.pk.to_python(company_id_str)
                selected_company = Company.objects.get(pk=company_pk, is_active=True)
                logger.info(
                    f"{log_prefix} SU using company '{selected_company.name}' (PK:{selected_company.pk}) from GET 'company_id'.")
                return selected_company
            except (Company.DoesNotExist, ValueError, TypeError) as e:
                logger.warning(f"{log_prefix} SU provided invalid company_id '{company_id_str}'. Error: {e}")
                raise Http404(
                    _("Company specified by ID '%(id)s' not found, invalid, or inactive.") % {'id': company_id_str})
        if isinstance(company_from_middleware, Company) and company_from_middleware.is_active:
            logger.info(
                f"{log_prefix} SU using company '{company_from_middleware.name}' (PK:{company_from_middleware.pk}) from request.company.")
            return company_from_middleware
        logger.info(f"{log_prefix} SU: No 'company_id' in GET and no active company from middleware. Prompting.")
        raise ValueError(
            _("Superuser: 'Company ID' parameter required or selected company is inactive. Please select an active company to view reports."))
    else:
        if not isinstance(company_from_middleware, Company) or not company_from_middleware.is_active:
            logger.warning(
                f"{log_prefix} Non-SU: No valid active company context (request.company is '{type(company_from_middleware)}').")
            raise DjangoPermissionDenied(
                _("No active company associated with your session, or company is inactive. Access denied."))
        logger.info(
            f"{log_prefix} Non-SU using company '{company_from_middleware.name}' (PK:{company_from_middleware.pk}) from request.company.")
        return company_from_middleware


def _default_link_callback(uri: str, rel: str) -> Optional[str]:
    """
    Default link callback for xhtml2pdf.
    Resolves local static and media files to absolute file system paths.
    """
    # Try to prevent issues if settings are not fully defined (e.g., during tests or partial setups)
    media_url = getattr(settings, 'MEDIA_URL', None)
    media_root = getattr(settings, 'MEDIA_ROOT', None)
    static_url = getattr(settings, 'STATIC_URL', None)
    static_root = getattr(settings, 'STATIC_ROOT', None)  # This is what collectstatic uses

    path = None
    if media_url and uri.startswith(media_url) and media_root:
        path = os.path.join(media_root, uri.replace(media_url, "", 1))
    elif static_url and uri.startswith(static_url) and static_root:  # Check static_root first
        path = os.path.join(static_root, uri.replace(static_url, "", 1))
    elif static_url and uri.startswith(
            static_url):  # Fallback for resolving against STATICFILES_DIRS if static_root not primary
        # This part is trickier as STATICFILES_DIRS can have multiple paths
        # For simplicity, we'll try finding it in the first one or rely on static_root
        # A more robust solution might iterate settings.STATICFILES_DIRS
        from django.contrib.staticfiles import finders
        # finders.find will search through STATICFILES_DIRS and app 'static' directories
        absolute_path = finders.find(uri.replace(static_url, "", 1))
        if absolute_path:
            path = absolute_path

    if not path:  # If not resolved yet, handle other cases
        if os.path.isabs(uri):  # If it's already an absolute path
            path = uri
        else:
            logger.debug(
                f"xhtml2pdf link_callback: URI '{uri}' is not a local static/media asset or absolute path, passing through.")
            return uri  # Let pisa try to handle it (e.g. http:// links)

    path = os.path.normpath(path)

    # Basic security check (ensure it's within known safe roots if resolved from URL)
    safe_roots = []
    if static_root: safe_roots.append(os.path.normpath(static_root))
    if media_root: safe_roots.append(os.path.normpath(media_root))
    # Also consider STATICFILES_DIRS if used above for resolution
    for staticfiles_dir_path in getattr(settings, 'STATICFILES_DIRS', []):
        safe_roots.append(os.path.normpath(staticfiles_dir_path))

    is_safe_path = False
    if any(path.startswith(safe_root) for safe_root in safe_roots):
        is_safe_path = True

    if not is_safe_path and not os.path.isabs(uri):  # If original uri was not absolute and path is not safe
        logger.warning(
            f"xhtml2pdf link_callback: URI '{uri}' resolved to potentially unsafe path '{path}'. Denying access.")
        return None

    if os.path.isfile(path):
        return path

    logger.warning(f"xhtml2pdf link_callback: file not found for URI '{uri}' (resolved to '{path}')")
    return None


# MODIFIED: Added 'request: Optional[HttpRequest] = None' parameter
def _render_to_pdf(template_src: str, context_dict: Optional[Dict[str, Any]] = None,
                   request: Optional[HttpRequest] = None) -> BytesIO:
    if not XHTML2PDF_AVAILABLE:
        logger.error("Attempted PDF generation but xhtml2pdf library is not installed.")
        raise ImportError(_("xhtml2pdf library is required for PDF generation. Please install it."))

    if context_dict is None:
        context_dict = {}

    # Diagnostic log for context_dict['request'] state if it exists from report_data
    if 'request' in context_dict:
        if not isinstance(context_dict['request'], HttpRequest):
            logger.warning(
                f"_render_to_pdf: context_dict already contains a 'request' key, but it's a "
                f"{type(context_dict['request']).__name__}, not HttpRequest. "
                f"This should be correctly handled by passing the 'request' parameter to render_to_string."
            )
        # If context_dict['request'] IS an HttpRequest, it's fine too.
        # render_to_string will prioritize the explicitly passed 'request' argument if different.

    current_link_callback = _default_link_callback
    required_settings_for_callback = ['STATIC_URL', 'MEDIA_URL']  # STATIC_ROOT/MEDIA_ROOT checked in callback
    if not all(hasattr(settings, s_name) and getattr(settings, s_name) for s_name in required_settings_for_callback):
        logger.warning(
            "xhtml2pdf: One or more of STATIC_URL, MEDIA_URL "  # STATIC_ROOT/MEDIA_ROOT are also important
            "are not configured in Django settings. Local file paths in PDF (images, CSS) via "
            "link_callback might not work as expected."
        )
        # Consider if link_callback should be disabled or simplified if settings are missing
        # current_link_callback = None # For example

    # MODIFIED: Pass the 'request' object to render_to_string
    html_content = render_to_string(template_src, context_dict, request=request)
    result_buffer = BytesIO()

    try:
        pdf_status = pisa.CreatePDF(
            BytesIO(html_content.encode("UTF-8")),
            dest=result_buffer,
            encoding='utf-8',
            link_callback=current_link_callback
        )
    except Exception as e:
        logger.exception(
            f"Exception during pisa.CreatePDF for template {template_src}. "
            f"HTML (first 500 chars): {html_content[:500]}"
        )
        raise Exception(
            _("PDF generation failed due to an internal error in the PDF library: %(error)s") % {'error': str(e)})

    if pdf_status.err:
        error_message = f"PDF generation error (code {pdf_status.err}) for template {template_src}."
        logger.error(f"{error_message} HTML (first 500 chars): {html_content[:500]}")
        # You can add more detailed pisa log messages here if needed
        # pisa_log = getattr(pdf_status, "log", []) # Check actual attribute name in pisa.document.Document
        # for msg_type, msg_val, msg_obj_id, msg_text in pisa_log:
        #     if msg_type >= logging.WARNING: # Example to log warnings and errors from pisa
        #         logger.error(f"PISA Log: Type={msg_type}, Value={msg_val}, Text={msg_text}")
        raise Exception(
            _("PDF generation failed (code %(code)s). Check server logs. "
              "This often indicates issues with HTML/CSS structure or unresolvable linked resources.") % {
                'code': pdf_status.err}
        )

    result_buffer.seek(0)
    return result_buffer


def _write_bs_hierarchy_excel(
        sheet: Any, nodes: List[Dict[str, Any]], start_row: int,
        level_offset: int = 0, currency_symbol: str = ""
) -> int:
    current_row = start_row
    bold_font = Font(bold=True)
    # Ensure currency_symbol is not None for formatting
    actual_currency_symbol = currency_symbol if currency_symbol is not None else ""
    standard_number_format = f'"{actual_currency_symbol}"#,##0.00;[Red]-"{actual_currency_symbol}"#,##0.00'
    for node in nodes:
        indentation = "    " * (node.get('level', 0) + level_offset)
        node_name = str(node.get('name', 'N/A'))
        name_cell_value = f"{indentation}{node_name}"
        name_cell = sheet.cell(row=current_row, column=1, value=name_cell_value)

        balance_value = node.get('balance', ZERO)
        balance_cell = sheet.cell(row=current_row, column=2, value=balance_value)
        balance_cell.number_format = standard_number_format
        balance_cell.alignment = Alignment(horizontal='right')

        if node.get('type') == 'group' or node.get('is_total', False):
            name_cell.font = bold_font
            balance_cell.font = bold_font

        current_row += 1
        child_nodes = node.get('children')
        if child_nodes:
            current_row = _write_bs_hierarchy_excel(sheet, child_nodes, current_row, level_offset + 1,
                                                    currency_symbol)  # Pass original symbol
    return current_row


def _get_validated_date_params_for_download(request: HttpRequest, date_param_names: List[str]) -> Dict[str, date]:
    dates: Dict[str, date] = {}
    for param_name in date_param_names:
        date_str = request.GET.get(param_name)
        if not date_str:
            raise ValueError(
                _("Missing required date parameter '%(param)s' for download.") % {'param': param_name})
        try:
            dates[param_name] = date.fromisoformat(date_str)
        except ValueError:
            raise ValueError(
                _("Invalid date format for parameter '%(param)s': %(value)s.") % {'param': param_name,
                                                                                  'value': date_str})
    return dates


# =========================================
# HTML Report Views (No changes here needed for this specific error)
# =========================================
@staff_member_required
def admin_trial_balance_view(request: HttpRequest) -> HttpResponse:
    context = _get_admin_base_context(str(_("Trial Balance Report")), request)
    target_company: Optional[Company] = None
    as_of_date_str = request.GET.get('as_of_date')

    context['as_of_date_param'] = as_of_date_str or date.today().isoformat()
    context['report_data_available'] = False

    try:
        target_company = _get_company_for_report_or_raise(request)
        context['company'] = target_company
        context['title'] = f"{str(_('Trial Balance'))} - {target_company.name}"

        if as_of_date_str:
            try:
                as_of_date_val = date.fromisoformat(as_of_date_str)
                context['as_of_date_to_display'] = as_of_date_val
            except ValueError:
                messages.error(request, _("Invalid date format for 'As of Date'."))
                return render(request, 'admin/crp_accounting/reports/trial_balance_report.html', context)

            report_data = reports_service.generate_trial_balance_structured(
                company_id=target_company.id, as_of_date=as_of_date_val,
                report_currency=getattr(target_company, 'default_currency_code', 'USD')
            )
            context.update(report_data)
            context['report_data_available'] = True
            if not report_data.get('is_balanced'):
                messages.warning(request, _("The Trial Balance is out of balance!"))
        elif 'company_id' in request.GET:
            messages.info(request, _("Please select an 'As of Date' to generate the Trial Balance."))

    except ValueError as ve:
        messages.error(request, str(ve))
        context['report_error'] = str(ve)
        if request.user.is_superuser and ("Company ID" in str(ve) or "select a company" in str(ve).lower()):
            if Company: context['all_companies'] = Company.objects.filter(is_active=True).order_by('name')
            context['show_company_selector'] = True
            context['title'] = str(_('Trial Balance - Select Company'))
    except (Http404, DjangoPermissionDenied) as e:
        messages.error(request, str(e))
        context['report_error'] = str(e)
    except ReportGenerationError as rge:
        messages.error(request, f"{str(_('Error generating report'))}: {rge}")
        context['report_error'] = f"{str(_('Error generating report'))}: {rge}"
    except Exception as e:
        logger.exception(f"Error generating admin TB for Co '{getattr(target_company, 'name', 'N/A')}'")
        messages.error(request, _("An unexpected error occurred."))
        context['report_error'] = str(_("An unexpected error occurred."))

    return render(request, 'admin/crp_accounting/reports/trial_balance_report.html', context)


@staff_member_required
def admin_profit_loss_view(request: HttpRequest) -> HttpResponse:
    context = _get_admin_base_context(str(_("Profit & Loss Statement")), request)
    target_company: Optional[Company] = None
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    today = date.today()
    context['start_date_param'] = start_date_str or today.replace(day=1).isoformat()
    context['end_date_param'] = end_date_str or today.isoformat()
    context['report_data_available'] = False

    try:
        target_company = _get_company_for_report_or_raise(request)
        context['company'] = target_company
        context['title'] = f"{str(_('Profit & Loss Statement'))} - {target_company.name}"

        if start_date_str and end_date_str:
            try:
                start_date_val = date.fromisoformat(start_date_str)
                end_date_val = date.fromisoformat(end_date_str)
                context['start_date_to_display'] = start_date_val
                context['end_date_to_display'] = end_date_val
            except ValueError:
                messages.error(request, _("Invalid date format for period."))
                return render(request, 'admin/crp_accounting/reports/profit_loss_report.html', context)

            if start_date_val > end_date_val:
                messages.error(request, _("Start date cannot be after end date."))
                return render(request, 'admin/crp_accounting/reports/profit_loss_report.html', context)

            report_data = reports_service.generate_profit_loss(
                company_id=target_company.id, start_date=start_date_val, end_date=end_date_val,
                report_currency=getattr(target_company, 'default_currency_code', 'USD')
            )
            context.update(report_data)
            context['report_data_available'] = True
        elif 'company_id' in request.GET:
            messages.info(request, _("Please select a 'From' and 'To' date to generate the P&L."))

    except ValueError as ve:
        messages.error(request, str(ve))
        context['report_error'] = str(ve)
        if request.user.is_superuser and ("Company ID" in str(ve) or "select a company" in str(ve).lower()):
            if Company: context['all_companies'] = Company.objects.filter(is_active=True).order_by('name')
            context['show_company_selector'] = True
            context['title'] = str(_('Profit & Loss - Select Company'))
    except (Http404, DjangoPermissionDenied) as e:
        messages.error(request, str(e))
        context['report_error'] = str(e)
    except ReportGenerationError as rge:
        messages.error(request, f"{str(_('Error generating report'))}: {rge}")
        context['report_error'] = f"{str(_('Error generating report'))}: {rge}"
    except Exception as e:
        logger.exception(f"Error generating admin P&L for Co '{getattr(target_company, 'name', 'N/A')}'")
        messages.error(request, _("An unexpected error occurred."))
        context['report_error'] = str(_("An unexpected error occurred."))

    return render(request, 'admin/crp_accounting/reports/profit_loss_report.html', context)


@staff_member_required
def admin_balance_sheet_view(request: HttpRequest) -> HttpResponse:
    context = _get_admin_base_context(str(_("Balance Sheet")), request)
    target_company: Optional[Company] = None
    as_of_date_str = request.GET.get('as_of_date')
    context['as_of_date_param'] = as_of_date_str or date.today().isoformat()
    context['report_data_available'] = False
    context['layout_preference'] = request.GET.get('layout', 'horizontal')
    if context['layout_preference'] not in ['horizontal', 'vertical']: context['layout_preference'] = 'horizontal'

    try:
        target_company = _get_company_for_report_or_raise(request)
        context['company'] = target_company
        context['title'] = f"{str(_('Balance Sheet'))} - {target_company.name}"

        if as_of_date_str:
            try:
                as_of_date_val = date.fromisoformat(as_of_date_str)
                context['as_of_date_to_display'] = as_of_date_val
            except ValueError:
                messages.error(request, _("Invalid date format for 'As of Date'."))
                return render(request, 'admin/crp_accounting/reports/balance_sheet_report.html', context)

            report_data = reports_service.generate_balance_sheet(
                company_id=target_company.id, as_of_date=as_of_date_val,
                report_currency=getattr(target_company, 'default_currency_code', 'USD')
            )
            context.update(report_data)
            context['report_data_available'] = True
            if not report_data.get('is_balanced'):
                messages.warning(request, _("The Balance Sheet is out of balance!"))
            if 'assets' in report_data and 'liabilities' in report_data and 'equity' in report_data:
                context['balance_difference'] = report_data['assets'].get('total', ZERO) - \
                                                (report_data['liabilities'].get('total', ZERO) + report_data[
                                                    'equity'].get('total', ZERO))
        elif 'company_id' in request.GET:
            messages.info(request, _("Please select an 'As of Date' to generate the Balance Sheet."))

    except ValueError as ve:
        messages.error(request, str(ve))
        context['report_error'] = str(ve)
        if request.user.is_superuser and ("Company ID" in str(ve) or "select a company" in str(ve).lower()):
            if Company: context['all_companies'] = Company.objects.filter(is_active=True).order_by('name')
            context['show_company_selector'] = True
            context['title'] = str(_('Balance Sheet - Select Company'))
    except (Http404, DjangoPermissionDenied) as e:
        messages.error(request, str(e))
        context['report_error'] = str(e)
    except ReportGenerationError as rge:
        messages.error(request, f"{str(_('Error generating report'))}: {rge}")
        context['report_error'] = f"{str(_('Error generating report'))}: {rge}"
    except Exception as e:
        logger.exception(f"Error generating admin BS for Co '{getattr(target_company, 'name', 'N/A')}'")
        messages.error(request, _("An unexpected error occurred."))
        context['report_error'] = str(_("An unexpected error occurred."))

    return render(request, 'admin/crp_accounting/reports/balance_sheet_report.html', context)


@staff_member_required
def admin_account_ledger_view(request: HttpRequest, account_pk: PK_TYPE) -> HttpResponse:
    context = _get_admin_base_context(str(_("Account Ledger")), request)
    target_company: Optional[Company] = None
    target_account: Optional[Account] = None
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    today = date.today()
    context['start_date_param'] = start_date_str or today.replace(day=1).isoformat()
    context['end_date_param'] = end_date_str or today.isoformat()
    context['report_data_available'] = False

    try:
        if request.user.is_superuser:
            company_id_from_get = request.GET.get('company_id')
            if company_id_from_get:
                target_company = get_object_or_404(Company, pk=company_id_from_get, is_active=True)
                target_account = get_object_or_404(Account.objects.select_related('company', 'account_group'),
                                                   pk=account_pk, company=target_company)
            else:
                target_account = get_object_or_404(Account.objects.select_related('company', 'account_group'),
                                                   pk=account_pk)
                if not target_account.company or not target_account.company.is_active:
                    raise Http404(_("Account is not associated with an active company, or company context is missing."))
                target_company = target_account.company
        else:
            company_from_request = getattr(request, 'company', None)
            if not isinstance(company_from_request, Company) or not company_from_request.is_active:
                raise DjangoPermissionDenied(
                    _("No active company associated with your session or company is inactive."))
            target_company = company_from_request
            target_account = get_object_or_404(Account.objects.select_related('company', 'account_group'),
                                               pk=account_pk, company=target_company)

        context['company'] = target_company
        context['account'] = target_account
        context[
            'title'] = f"{str(_('Ledger'))}: {target_account.account_name} ({target_account.account_number}) - {target_company.name}"

        if start_date_str and end_date_str:
            try:
                start_date_val = date.fromisoformat(start_date_str)
                end_date_val = date.fromisoformat(end_date_str)
                context['start_date_to_display'] = start_date_val
                context['end_date_to_display'] = end_date_val
            except ValueError:
                messages.error(request, _("Invalid date format for ledger period."))
                return render(request, 'admin/crp_accounting/reports/account_ledger_report.html', context)

            if start_date_val > end_date_val:
                messages.error(request, _("Start date cannot be after end date."))
                return render(request, 'admin/crp_accounting/reports/account_ledger_report.html', context)

            ledger_data = ledger_service.get_account_ledger_data(
                company_id=target_company.id, account_pk=target_account.pk,
                start_date=start_date_val, end_date=end_date_val
            )
            is_debit_nature = target_account.account_nature == AccountNature.DEBIT.value

            def _format_balance_display(bal: Decimal) -> Dict[str, Any]:
                amt = abs(bal)
                dr_cr_indicator = ""
                if bal != ZERO:
                    dr_cr_indicator = ("Dr" if bal > ZERO else "Cr") if is_debit_nature else \
                        ("Cr" if bal > ZERO else "Dr")
                return {'amount': amt, 'dr_cr': dr_cr_indicator}

            context['opening_balance_display'] = _format_balance_display(ledger_data.get('opening_balance', ZERO))
            context['closing_balance_display'] = _format_balance_display(ledger_data.get('closing_balance', ZERO))
            processed_entries = []
            running_bal_for_display = ledger_data.get('opening_balance', ZERO)
            for entry in ledger_data.get('entries', []):
                debit, credit = entry.get('debit', ZERO), entry.get('credit', ZERO)
                if is_debit_nature:
                    running_bal_for_display += (debit - credit)
                else:
                    running_bal_for_display += (credit - debit)

                vt_label = str(VoucherType(entry['voucher_type']).label) if entry.get('voucher_type') and entry[
                    'voucher_type'] in VoucherType._value2member_map_ else ""
                processed_entries.append(
                    {**entry, 'running_balance_display': _format_balance_display(running_bal_for_display),
                     'voucher_type_display': vt_label})
            context['entries'] = processed_entries
            context['total_debit'] = ledger_data.get('total_debit', ZERO)
            context['total_credit'] = ledger_data.get('total_credit', ZERO)
            context['report_currency'] = target_account.currency
            context['report_data_available'] = True
        elif context.get('company') and context.get('account'):
            messages.info(request, _("Please select a 'From' and 'To' date to view the ledger details."))

    except DjangoPermissionDenied as e:
        messages.error(request, str(e))
        context['report_error'] = str(e)
    except Http404:
        messages.error(request,
                       _("Account or Company not found for ledger. Ensure company context is correct if superuser."))
        context['report_error'] = _("Account not found for ledger.")
    except ReportGenerationError as rge:
        messages.error(request, f"{str(_('Ledger error'))}: {rge}")
        context[
            'report_error'] = f"{str(_('Ledger error'))}: {rge}"
    except Exception as e:
        logger.exception(f"Unexpected error generating Account Ledger for AccPK {account_pk}")
        messages.error(request, _("An unexpected error occurred."))
        context['report_error'] = _("An unexpected error occurred.")
    return render(request, 'admin/crp_accounting/reports/account_ledger_report.html', context)


@staff_member_required
def admin_ar_aging_report_view(request: HttpRequest) -> HttpResponse:
    context = _get_admin_base_context(str(_("AR Aging Report")), request)
    target_company: Optional[Company] = None
    today = date.today()

    context['as_of_date_param'] = request.GET.get('as_of_date', today.isoformat())
    context['aging_bucket_days_config_str'] = ",".join(
        map(str, reports_service.DEFAULT_AR_AGING_BUCKETS_DAYS))
    context['report_data_available'] = False

    try:
        target_company = _get_company_for_report_or_raise(request)
        context['company'] = target_company
        context['title'] = f"{str(_('AR Aging Report'))} - {target_company.name}"

        as_of_date_val = date.fromisoformat(context['as_of_date_param'])

        buckets_str_from_get = request.GET.get('aging_buckets')
        if buckets_str_from_get:
            try:
                aging_buckets_to_use = [int(b.strip()) for b in buckets_str_from_get.split(',') if b.strip()]
                context['aging_bucket_days_config_str'] = buckets_str_from_get
            except ValueError:
                messages.error(request, _("Invalid format for aging buckets. Using default."))
                aging_buckets_to_use = reports_service.DEFAULT_AR_AGING_BUCKETS_DAYS
        else:
            aging_buckets_to_use = reports_service.DEFAULT_AR_AGING_BUCKETS_DAYS

        report_data = reports_service.generate_ar_aging_report(
            company_id=target_company.id,
            as_of_date=as_of_date_val,
            report_currency=getattr(target_company, 'default_currency_code', 'USD'),
            aging_buckets_days=aging_buckets_to_use
        )
        context.update(report_data)
        context['report_data_available'] = True

    except ValueError as ve:
        messages.error(request, str(ve))
        context['report_error'] = str(ve)
        if "Company ID" in str(ve) and request.user.is_superuser:
            if Company: context['all_companies'] = Company.objects.filter(is_active=True).order_by('name')
            context['show_company_selector'] = True
            context['title'] = str(_('AR Aging Report - Select Company'))
    except (Http404, DjangoPermissionDenied) as e:
        messages.error(request, str(e))
        context['report_error'] = str(e)
    except ReportGenerationError as rge:
        messages.error(request, f"{str(_('Error generating AR Aging report'))}: {rge}")
        context['report_error'] = f"{str(_('Error generating AR Aging report'))}: {rge}"
    except Exception as e:
        logger.exception(f"Error generating admin AR Aging Report for Co '{getattr(target_company, 'name', 'N/A')}'")
        messages.error(request, _("An unexpected error occurred while generating the AR Aging report."))
        context['report_error'] = str(_("An unexpected error occurred."))

    if 'opts' not in context: context['opts'] = Company._meta if Company else Account._meta

    return render(request, 'admin/crp_accounting/reports/ar_aging_report.html', context)


@staff_member_required
def admin_customer_statement_view(request: HttpRequest, customer_pk: Optional[PK_TYPE] = None) -> HttpResponse:
    context = _get_admin_base_context(str(_("Customer Statement")), request)
    target_company: Optional[Company] = None
    target_customer: Optional[Party] = None
    today = date.today()
    default_start_date = today.replace(day=1)
    next_month = default_start_date.replace(day=28) + timedelta(days=4)
    default_end_date = next_month - timedelta(days=next_month.day)

    context['start_date_param'] = request.GET.get('start_date', default_start_date.isoformat())
    context['end_date_param'] = request.GET.get('end_date', default_end_date.isoformat())
    customer_pk_from_get = request.GET.get('customer_id')
    effective_customer_pk = customer_pk or customer_pk_from_get
    context['report_data_available'] = False

    try:
        target_company = _get_company_for_report_or_raise(request)
        context['company'] = target_company
        context['title'] = f"{str(_('Customer Statement'))} - {target_company.name}"

        if target_company:
            context['customers_for_selection'] = Party.objects.filter(
                company=target_company,
                party_type=CorePartyType.CUSTOMER.value,
                is_active=True
            ).order_by('name')

        if not effective_customer_pk:
            if 'company_id' in request.GET and not customer_pk_from_get:
                messages.info(request, _("Please select a customer to view the statement."))
                context['show_customer_selector'] = True
        else:
            target_customer = get_object_or_404(
                Party.objects.select_related('company'),
                pk=effective_customer_pk,
                company=target_company,
                party_type=CorePartyType.CUSTOMER.value
            )
            context['customer'] = target_customer
            context['title'] = f"{str(_('Statement'))}: {target_customer.name} - {target_company.name}"

            start_date_val = date.fromisoformat(context['start_date_param'])
            end_date_val = date.fromisoformat(context['end_date_param'])
            if start_date_val > end_date_val:
                raise DjangoValidationError(_("Start date cannot be after end date for the statement."))
            context['start_date_to_display'] = start_date_val
            context['end_date_to_display'] = end_date_val

            report_data = reports_service.generate_customer_statement(
                company_id=target_company.id,
                customer_id=target_customer.pk,
                start_date=start_date_val,
                end_date=end_date_val,
                report_currency=getattr(target_company, 'default_currency_code', 'USD')
            )
            context.update(report_data)
            context['report_data_available'] = True

    except ValueError as ve:
        messages.error(request, str(ve))
        context['report_error'] = str(ve)
        if "Company ID" in str(ve) and request.user.is_superuser and not target_company:
            if Company: context['all_companies'] = Company.objects.filter(is_active=True).order_by('name')
            context['show_company_selector'] = True
            context['title'] = str(_('Customer Statement - Select Company'))
    except (Http404, DjangoPermissionDenied) as e:
        messages.error(request, str(e))
        context['report_error'] = str(e)
    except DjangoValidationError as dve:
        error_message = "; ".join(dve.messages) if hasattr(dve, 'messages') and isinstance(dve.messages, list) else str(
            dve)
        messages.error(request, error_message)
        context['report_error'] = error_message
    except ReportGenerationError as rge:
        messages.error(request, f"{str(_('Error generating statement'))}: {rge}")
        context['report_error'] = f"{str(_('Error generating statement'))}: {rge}"
    except Exception as e:
        logger.exception(
            f"Error generating Customer Statement for Co '{getattr(target_company, 'name', 'N/A')}', Cust '{getattr(target_customer, 'name', effective_customer_pk)}'")
        messages.error(request, _("An unexpected error occurred generating the statement."))
        context['report_error'] = str(_("An unexpected error occurred."))

    if 'opts' not in context: context['opts'] = Party._meta

    return render(request, 'admin/crp_accounting/reports/customer_statement_report.html', context)


# =========================================
# EXCEL Download Views (CORRECTIONS APPLIED HERE)
# =========================================
@staff_member_required
def download_trial_balance_excel(request: HttpRequest) -> HttpResponse:
    target_company: Optional[Company] = None
    if not OPENPYXL_AVAILABLE:
        return HttpResponse(str(_("Excel export library (openpyxl) is not installed.")), status=501)
    try:
        target_company = _get_company_for_report_or_raise(request)
        date_params = _get_validated_date_params_for_download(request, ['as_of_date'])
        as_of_date_val = date_params['as_of_date']

        report_data = reports_service.generate_trial_balance_structured(
            company_id=target_company.id, as_of_date=as_of_date_val,
            report_currency=getattr(target_company, 'default_currency_code', 'USD')
        )
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = str(_("Trial Balance"))

        # CORRECTED currency_symbol logic
        currency_symbol_candidate = getattr(target_company, 'default_currency_symbol', None)
        currency_symbol = currency_symbol_candidate if currency_symbol_candidate is not None else ''

        sheet['A1'] = f"{str(_('Trial Balance Report'))} - {target_company.name}"
        sheet.merge_cells('A1:D1')
        sheet['A1'].font = Font(bold=True, size=14)
        sheet['A1'].alignment = Alignment(horizontal='center')
        sheet['A2'] = f"{str(_('As of Date:'))} {as_of_date_val.strftime('%B %d, %Y')}"
        sheet.merge_cells('A2:D2')
        sheet['A2'].alignment = Alignment(horizontal='center')
        sheet.append([])
        headers = [str(_("Account Number")), str(_("Account Name")), str(_("Debit")), str(_("Credit"))]
        sheet.append(headers)
        header_row_num = sheet.max_row
        for col_idx, cell_value in enumerate(headers, 1):
            cell = sheet.cell(row=header_row_num, column=col_idx)
            cell.font = Font(bold=True)
            cell.border = Border(bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal='center' if col_idx != 2 else 'left')
            sheet.column_dimensions[get_column_letter(col_idx)].width = (
                20 if col_idx == 1 else 35 if col_idx == 2 else 18)
        sheet.cell(row=header_row_num, column=3).alignment = Alignment(horizontal='right')
        sheet.cell(row=header_row_num, column=4).alignment = Alignment(horizontal='right')
        number_format_currency = f'"{currency_symbol}"#,##0.00;[Red]-"{currency_symbol}"#,##0.00'

        for entry in report_data.get('flat_entries', []):
            if entry.get('debit', ZERO) == ZERO and entry.get('credit', ZERO) == ZERO and not entry.get('is_group'):
                continue
            row_values = [entry.get('account_number', ''), str(entry.get('account_name', 'N/A')),
                          entry.get('debit', ZERO), entry.get('credit', ZERO)]
            sheet.append(row_values)
            current_row = sheet.max_row
            sheet.cell(row=current_row, column=3).number_format = number_format_currency
            sheet.cell(row=current_row, column=4).number_format = number_format_currency
            if entry.get('is_group'):
                for col_idx_group in range(1, 5):
                    sheet.cell(row=current_row, column=col_idx_group).font = Font(bold=True)

        sheet.append([])
        sheet.append(
            [None, str(_("Totals")), report_data.get('total_debit', ZERO), report_data.get('total_credit', ZERO)])
        total_row_num = sheet.max_row
        for col_idx_total in [2, 3, 4]:
            sheet.cell(row=total_row_num, column=col_idx_total).font = Font(bold=True)
        sheet.cell(row=total_row_num, column=3).number_format = number_format_currency
        sheet.cell(row=total_row_num, column=4).number_format = number_format_currency
        sheet.cell(row=total_row_num, column=3).alignment = Alignment(horizontal='right')
        sheet.cell(row=total_row_num, column=4).alignment = Alignment(horizontal='right')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"{target_company.name}_Trial_Balance_{as_of_date_val.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response
    except ValueError as ve:
        return HttpResponseBadRequest(str(ve))
    except (Http404, DjangoPermissionDenied) as e:
        return HttpResponse(str(e), status=403 if isinstance(e, DjangoPermissionDenied) else 404)
    except ReportGenerationError as rge:
        return HttpResponse(f"{str(_('Report generation error'))}: {rge}", status=500)
    except Exception as e:
        logger.exception(f"Error exporting Trial Balance Excel for Co '{getattr(target_company, 'name', 'N/A')}'")
        return HttpResponse(str(_("An unexpected error occurred during Excel generation.")), status=500)


@staff_member_required
def download_profit_loss_excel(request: HttpRequest) -> HttpResponse:
    target_company: Optional[Company] = None
    if not OPENPYXL_AVAILABLE:
        return HttpResponse(str(_("Excel export library (openpyxl) is not installed.")), status=501)
    try:
        target_company = _get_company_for_report_or_raise(request)
        date_params = _get_validated_date_params_for_download(request, ['start_date', 'end_date'])
        start_date_val, end_date_val = date_params['start_date'], date_params['end_date']
        if start_date_val > end_date_val: raise ValueError(str(_("Start date cannot be after end date.")))

        report_data = reports_service.generate_profit_loss(
            company_id=target_company.id, start_date=start_date_val, end_date=end_date_val,
            report_currency=getattr(target_company, 'default_currency_code', 'USD')
        )
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = str(_("Profit & Loss"))

        # CORRECTED currency_symbol logic
        currency_symbol_candidate = getattr(target_company, 'default_currency_symbol', None)
        currency_symbol = currency_symbol_candidate if currency_symbol_candidate is not None else ''

        number_format_currency = f'"{currency_symbol}"#,##0.00;[Red]-"{currency_symbol}"#,##0.00'
        sheet['A1'] = f"{str(_('Profit & Loss Statement'))} - {target_company.name}"
        sheet.merge_cells('A1:C1')
        sheet['A1'].font = Font(bold=True, size=14)
        sheet['A1'].alignment = Alignment(horizontal='center')
        sheet[
            'A2'] = f"{str(_('For the period:'))} {start_date_val.strftime('%B %d, %Y')} {str(_('to'))} {end_date_val.strftime('%B %d, %Y')}"
        sheet.merge_cells('A2:C2')
        sheet['A2'].alignment = Alignment(horizontal='center')
        sheet.append([])
        headers = [str(_("Description")), str(_("Amount")), str(_("Details"))]
        sheet.append(headers)
        header_row_num = sheet.max_row
        for col_idx, cell_val_pl in enumerate(headers, 1):
            cell = sheet.cell(row=header_row_num, column=col_idx)
            cell.font = Font(bold=True)
            cell.border = Border(bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal='left' if col_idx == 1 else 'right')
        sheet.column_dimensions[get_column_letter(1)].width = 45
        sheet.column_dimensions[get_column_letter(2)].width = 20
        sheet.column_dimensions[get_column_letter(3)].width = 30

        for item in report_data.get('report_lines', []):
            is_subtotal_or_net = item.get('is_subtotal', False) or item.get('section_key') == 'NET_INCOME'
            row_data = [str(item.get('title', 'N/A')), item.get('amount', ZERO)]
            account_details_str = ""
            if not item.get('is_subtotal', False) and item.get('accounts'):
                if len(item['accounts']) <= 3:
                    account_details_str = "; ".join(
                        [
                            f"{acc['account_number']} - {acc['account_name']}: {currency_symbol}{acc.get('amount', ZERO):,.2f}"
                            for acc in
                            item['accounts']])
                elif item.get('has_note') and item.get('note_ref'):
                    account_details_str = f"See Note: {item['note_ref']}"
            row_data.append(account_details_str if account_details_str else None)
            sheet.append(row_data)

            current_row_num = sheet.max_row
            desc_cell = sheet.cell(row=current_row_num, column=1)
            amt_cell = sheet.cell(row=current_row_num, column=2)
            desc_cell.font = Font(bold=is_subtotal_or_net)
            amt_cell.font = Font(bold=is_subtotal_or_net)
            amt_cell.number_format = number_format_currency

        if report_data.get('financial_notes_data'):
            sheet.append([])
            sheet.append([str(_("Financial Notes"))])
            sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True, size=12)
            for note_ref, note_content in report_data['financial_notes_data'].items():
                sheet.append([f"{str(note_content.get('title', note_ref))} ({note_ref})"])
                sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True)
                for detail_line in note_content.get('details', []):
                    note_desc = detail_line.get('account_name', detail_line.get('title', ''))
                    note_amt = detail_line.get('amount', ZERO)
                    sheet.append([f"    {str(note_desc)}", note_amt])
                    sheet.cell(row=sheet.max_row, column=2).number_format = number_format_currency
                sheet.append([str(_("Total for Note")) + f" {note_ref}", note_content.get('total_amount', ZERO)])
                total_note_cell = sheet.cell(row=sheet.max_row, column=1)
                total_note_amt_cell = sheet.cell(row=sheet.max_row, column=2)
                total_note_cell.font = Font(italic=True)
                total_note_amt_cell.font = Font(italic=True)
                total_note_amt_cell.number_format = number_format_currency

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"{target_company.name}_Profit_Loss_{start_date_val.strftime('%Y%m%d')}_{end_date_val.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response
    except ValueError as ve:
        return HttpResponseBadRequest(str(ve))
    except (Http404, DjangoPermissionDenied) as e:
        return HttpResponse(str(e), status=403 if isinstance(e, DjangoPermissionDenied) else 404)
    except ReportGenerationError as rge:
        return HttpResponse(f"{str(_('Report generation error'))}: {rge}", status=500)
    except Exception as e:
        logger.exception(f"Error exporting P&L Excel for Co '{getattr(target_company, 'name', 'N/A')}'")
        return HttpResponse(str(_("Excel generation error.")), status=500)


@staff_member_required
def download_balance_sheet_excel(request: HttpRequest) -> HttpResponse:
    target_company: Optional[Company] = None
    if not OPENPYXL_AVAILABLE:
        return HttpResponse(str(_("Excel export library (openpyxl) is not installed.")), status=501)
    try:
        target_company = _get_company_for_report_or_raise(request)
        date_params = _get_validated_date_params_for_download(request, ['as_of_date'])
        as_of_date_val = date_params['as_of_date']
        report_data = reports_service.generate_balance_sheet(
            company_id=target_company.id, as_of_date=as_of_date_val,
            report_currency=getattr(target_company, 'default_currency_code', 'USD')
        )
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = str(_("Balance Sheet"))

        # CORRECTED currency_symbol logic
        currency_symbol_candidate = getattr(target_company, 'default_currency_symbol', None)
        currency_symbol = currency_symbol_candidate if currency_symbol_candidate is not None else ''

        sheet['A1'] = f"{str(_('Balance Sheet'))} - {target_company.name}"
        sheet.merge_cells('A1:B1')
        sheet['A1'].font = Font(bold=True, size=14)
        sheet['A1'].alignment = Alignment(horizontal='center')
        sheet['A2'] = f"{str(_('As of Date:'))} {as_of_date_val.strftime('%B %d, %Y')}"
        sheet.merge_cells('A2:B2')
        sheet['A2'].alignment = Alignment(horizontal='center')
        sheet.column_dimensions['A'].width = 50
        sheet.column_dimensions['B'].width = 20
        current_row = 3

        number_format_currency_bs = f'"{currency_symbol}"#,##0.00;[Red]-"{currency_symbol}"#,##0.00'

        current_row += 1
        sheet.cell(row=current_row, column=1, value=str(_("Assets"))).font = Font(bold=True, size=12)
        current_row += 1
        assets_data = report_data.get('assets', {})
        current_row = _write_bs_hierarchy_excel(sheet, assets_data.get('hierarchy', []), current_row, 0,
                                                currency_symbol)  # currency_symbol is now corrected
        sheet.cell(row=current_row, column=1, value=str(_("Total Assets"))).font = Font(bold=True)
        total_assets_cell = sheet.cell(row=current_row, column=2, value=assets_data.get('total', ZERO))
        total_assets_cell.font = Font(bold=True)
        total_assets_cell.number_format = number_format_currency_bs
        total_assets_cell.alignment = Alignment(horizontal='right')
        current_row += 2

        sheet.cell(row=current_row, column=1, value=str(_("Liabilities"))).font = Font(bold=True, size=12)
        current_row += 1
        liabilities_data = report_data.get('liabilities', {})
        current_row = _write_bs_hierarchy_excel(sheet, liabilities_data.get('hierarchy', []), current_row, 0,
                                                currency_symbol)  # currency_symbol is now corrected
        sheet.cell(row=current_row, column=1, value=str(_("Total Liabilities"))).font = Font(bold=True)
        total_liab_cell = sheet.cell(row=current_row, column=2, value=liabilities_data.get('total', ZERO))
        total_liab_cell.font = Font(bold=True)
        total_liab_cell.number_format = number_format_currency_bs
        total_liab_cell.alignment = Alignment(horizontal='right')
        current_row += 2

        sheet.cell(row=current_row, column=1, value=str(_("Equity"))).font = Font(bold=True, size=12)
        current_row += 1
        equity_data = report_data.get('equity', {})
        current_row = _write_bs_hierarchy_excel(sheet, equity_data.get('hierarchy', []), current_row, 0,
                                                currency_symbol)  # currency_symbol is now corrected
        sheet.cell(row=current_row, column=1, value=str(_("Total Equity"))).font = Font(bold=True)
        total_equity_cell = sheet.cell(row=current_row, column=2, value=equity_data.get('total', ZERO))
        total_equity_cell.font = Font(bold=True)
        total_equity_cell.number_format = number_format_currency_bs
        total_equity_cell.alignment = Alignment(horizontal='right')
        current_row += 2

        total_liab_equity = liabilities_data.get('total', ZERO) + equity_data.get('total', ZERO)
        sheet.cell(row=current_row, column=1, value=str(_("Total Liabilities and Equity"))).font = Font(bold=True)
        total_liab_equity_cell = sheet.cell(row=current_row, column=2, value=total_liab_equity)
        total_liab_equity_cell.font = Font(bold=True)
        total_liab_equity_cell.number_format = number_format_currency_bs
        total_liab_equity_cell.alignment = Alignment(horizontal='right')

        if not report_data.get('is_balanced', True):
            current_row += 2
            warning_cell = sheet.cell(row=current_row, column=1, value=str(_("Note: Balance Sheet is Out of Balance!")))
            warning_cell.font = Font(color="FF0000", bold=True)
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
            warning_cell.alignment = Alignment(horizontal='center')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"{target_company.name}_Balance_Sheet_{as_of_date_val.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response
    except ValueError as ve:
        return HttpResponseBadRequest(str(ve))
    except (Http404, DjangoPermissionDenied) as e:
        return HttpResponse(str(e), status=403 if isinstance(e, DjangoPermissionDenied) else 404)
    except ReportGenerationError as rge:
        return HttpResponse(f"{str(_('Report generation error'))}: {rge}", status=500)
    except Exception as e:
        logger.exception(f"Error exporting BS Excel for Co '{getattr(target_company, 'name', 'N/A')}'")
        return HttpResponse(str(_("Excel generation error.")), status=500)

@staff_member_required
def download_ar_aging_excel(request: HttpRequest) -> HttpResponse:
    target_company: Optional[Company] = None
    if not OPENPYXL_AVAILABLE:
        return HttpResponse(str(_("Excel export library (openpyxl) is not installed.")), status=501)

    try:
        target_company = _get_company_for_report_or_raise(request)
        date_params = _get_validated_date_params_for_download(request, ['as_of_date'])
        as_of_date_val = date_params['as_of_date']

        buckets_str_from_get = request.GET.get('aging_buckets')
        if buckets_str_from_get:
            try:
                aging_buckets_to_use = [int(b.strip()) for b in buckets_str_from_get.split(',') if b.strip()]
            except ValueError:
                aging_buckets_to_use = reports_service.DEFAULT_AR_AGING_BUCKETS_DAYS
                logger.warning(
                    f"Invalid aging_buckets GET param '{buckets_str_from_get}'. Using default for Excel export.")
        else:
            aging_buckets_to_use = reports_service.DEFAULT_AR_AGING_BUCKETS_DAYS

        report_data = reports_service.generate_ar_aging_report(
            company_id=target_company.id,
            as_of_date=as_of_date_val,
            report_currency=getattr(target_company, 'default_currency_code', 'USD'),
            aging_buckets_days=aging_buckets_to_use
        )

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = str(_("AR Aging"))

        # CORRECTED currency_symbol logic
        company_sym = getattr(target_company, 'default_currency_symbol', None)
        report_sym = report_data.get('report_currency', None)
        if company_sym is not None:
            currency_symbol = company_sym
        elif report_sym is not None:
            currency_symbol = report_sym
        else:
            currency_symbol = ''

        number_format_currency = f'"{currency_symbol}"#,##0.00;[Red]-"{currency_symbol}"#,##0.00'

        sheet['A1'] = f"{str(_('AR Aging Report'))} - {target_company.name}"
        sheet.merge_cells(start_row=1, start_column=1, end_row=1,
                          end_column=len(report_data.get('bucket_labels', [])) + 2)
        sheet['A1'].font = Font(bold=True, size=14)
        sheet['A1'].alignment = Alignment(horizontal='center')

        sheet['A2'] = f"{str(_('As of Date:'))} {as_of_date_val.strftime('%B %d, %Y')}"
        sheet.merge_cells(start_row=2, start_column=1, end_row=2,
                          end_column=len(report_data.get('bucket_labels', [])) + 2)
        sheet['A2'].alignment = Alignment(horizontal='center')
        sheet.append([])

        excel_headers = [str(_("Customer Name"))] + [str(label) for label in report_data.get('bucket_labels', [])] + [
            str(_("Total Due"))]
        sheet.append(excel_headers)
        header_row_num = sheet.max_row
        for col_idx, header_title in enumerate(excel_headers, 1):
            cell = sheet.cell(row=header_row_num, column=col_idx, value=header_title)
            cell.font = Font(bold=True)
            cell.border = Border(bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal='left' if col_idx == 1 else 'right')
            sheet.column_dimensions[get_column_letter(col_idx)].width = 30 if col_idx == 1 else 18

        for entry in report_data.get('aging_data', []):
            row_values = [str(entry.get('customer_name', 'N/A'))]
            for bucket_label in report_data.get('bucket_labels', []):
                row_values.append(entry.get('buckets', {}).get(bucket_label, ZERO))
            row_values.append(entry.get('total_due', ZERO))
            sheet.append(row_values)
            current_data_row = sheet.max_row
            for col_idx in range(2, len(excel_headers) + 1):
                sheet.cell(row=current_data_row, column=col_idx).number_format = number_format_currency
                sheet.cell(row=current_data_row, column=col_idx).alignment = Alignment(horizontal='right')

        sheet.append([])

        grand_total_row_values = [str(_("Grand Totals"))]
        for bucket_label in report_data.get('bucket_labels', []):
            grand_total_row_values.append(report_data.get('grand_totals_by_bucket', {}).get(bucket_label, ZERO))
        grand_total_row_values.append(report_data.get('grand_total_due_all_customers', ZERO))
        sheet.append(grand_total_row_values)
        total_row_num = sheet.max_row
        for col_idx in range(1, len(grand_total_row_values) + 1):
            cell = sheet.cell(row=total_row_num, column=col_idx)
            cell.font = Font(bold=True)
            if col_idx > 1:
                cell.number_format = number_format_currency
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"{target_company.name}_AR_Aging_{as_of_date_val.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response

    except ValueError as ve:
        return HttpResponseBadRequest(str(ve))
    except (Http404, DjangoPermissionDenied) as e:
        return HttpResponse(str(e), status=403 if isinstance(e, DjangoPermissionDenied) else 404)
    except ReportGenerationError as rge:
        return HttpResponse(f"{str(_('Report generation error'))}: {rge}", status=500)
    except Exception as e:
        logger.exception(f"Error exporting AR Aging Excel for Co '{getattr(target_company, 'name', 'N/A')}'")
        return HttpResponse(str(_("An unexpected error occurred during AR Aging Excel generation.")), status=500)


@staff_member_required
def download_customer_statement_excel(request: HttpRequest, customer_pk: str) -> HttpResponse:
    """
    Generates and serves a Customer Statement as a styled Excel file.
    This version fixes column widths and data placement.
    """
    if not OPENPYXL_AVAILABLE:
        return HttpResponse(str(_("Excel export library (openpyxl) is not installed.")), status=501)

    target_company: Optional[Company] = None
    target_customer: Optional[Party] = None
    try:
        company_id_from_get = request.GET.get('company_id')
        if not company_id_from_get:
            raise ValueError("Company ID is required for download.")

        target_company = get_object_or_404(Company, pk=company_id_from_get)
        date_params = _get_validated_date_params_for_download(request, ['start_date', 'end_date'])
        start_date_val, end_date_val = date_params['start_date'], date_params['end_date']

        target_customer = get_object_or_404(
            Party.objects.select_related('company'),
            pk=customer_pk,
            company=target_company,
            party_type=CorePartyType.CUSTOMER.value
        )

        report_data = reports_service.generate_customer_statement(
            company_id=target_company.id,
            customer_id=target_customer.pk,
            start_date=start_date_val,
            end_date=end_date_val,
            report_currency=getattr(target_company, 'default_currency_code', 'USD')
        )

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = str(_("Customer Statement"))

        # --- STYLING SETUP ---
        currency_symbol = report_data.get('report_currency_symbol', '$')
        number_format_currency = f'"{currency_symbol}" #,##0.00'
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="004C99", end_color="004C99", fill_type="solid")
        title_font = Font(bold=True, size=14)
        center_align = Alignment(horizontal='center')
        bold_font = Font(bold=True)

        # --- WRITE TITLES ---
        sheet['A1'] = f"{str(_('Statement For:'))} {target_customer.name}"
        sheet.merge_cells('A1:G1');
        sheet['A1'].font = title_font
        sheet['A2'] = f"{str(_('Company:'))} {target_company.name}"
        sheet.merge_cells('A2:G2')
        sheet[
            'A3'] = f"{str(_('Period:'))} {start_date_val.strftime('%d-%b-%Y')} to {end_date_val.strftime('%d-%b-%Y')}"
        sheet.merge_cells('A3:G3')
        sheet.append([])

        # --- WRITE TABLE HEADERS ---
        headers = ["Date", "Transaction Type", "Reference", "Debit", "Credit", "Balance", "Dr/Cr"]
        sheet.append([str(_(h)) for h in headers])
        header_row_num = sheet.max_row
        for cell in sheet[header_row_num]:
            cell.font = header_font;
            cell.fill = header_fill

        # --- WRITE DATA ROWS (Corrected to match web page structure) ---
        ob = report_data.get('opening_balance', ZERO)
        ob_drcr = 'Dr' if ob > 0 else ('Cr' if ob < 0 else '')
        # Opening Balance Row
        sheet.append([
            start_date_val.strftime('%Y-%m-%d'),
            str(_("Opening Balance")),
            "",  # Reference is blank
            None,  # Debit is blank
            None,  # Credit is blank
            ob,  # Balance
            ob_drcr  # Dr/Cr
        ])

        # Transaction Rows
        for line in report_data.get('lines', []):
            balance = line.get('balance', ZERO)
            drcr_indicator = 'Dr' if balance > 0 else ('Cr' if balance < 0 else '')
            sheet.append([
                line.get('date'),
                line.get('transaction_type'),
                line.get('reference'),
                line.get('debit'),
                line.get('credit'),
                balance,
                drcr_indicator
            ])

        # Closing Balance Row
        cb = report_data.get('closing_balance', ZERO)
        cb_drcr = 'Dr' if cb > 0 else ('Cr' if cb < 0 else '')
        sheet.append([
            end_date_val.strftime('%Y-%m-%d'),
            str(_("Closing Balance")),
            "", None, None,
            cb,
            cb_drcr
        ])

        # --- APPLY FORMATTING & BOLDING ---
        data_start_row = header_row_num + 1
        for row in sheet.iter_rows(min_row=data_start_row, max_row=sheet.max_row):
            row[0].number_format = 'YYYY-MM-DD'  # Date column (A)
            row[3].number_format = number_format_currency  # Debit column (D)
            row[4].number_format = number_format_currency  # Credit column (E)
            row[5].number_format = number_format_currency  # Balance column (F)
            row[6].alignment = center_align  # Dr/Cr column (G)

        # Bold the opening and closing balance rows
        for cell in sheet[data_start_row]: cell.font = bold_font
        for cell in sheet[sheet.max_row]: cell.font = bold_font

        # --- NEW: AUTO-SIZE COLUMNS ---
        for col_idx in range(1, sheet.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            # Find the longest cell value in the column
            for cell in sheet[column_letter]:
                try:
                    # Check if the cell has a value and is not a merged cell
                    if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            # Add some padding to the width
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

        # --- CREATE AND RETURN HTTP RESPONSE ---
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Customer_Statement_{target_customer.name.replace(' ', '_')}_{end_date_val.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response

    except (ValueError, Http404) as e:
        return HttpResponseBadRequest(str(e))
    except Exception as e:
        logger.exception(
            f"Error generating Customer Statement Excel for Cust '{getattr(target_customer, 'name', customer_pk)}'")
        return HttpResponse(str(_("An unexpected error occurred during Excel generation.")), status=500)
# =============================================================================
# NEW: Accounts Payable (AP) HTML Report Views (Tenant Aware)
# =============================================================================
@staff_member_required
def admin_ap_aging_report_view(request: HttpRequest) -> HttpResponse:
    """Displays the Accounts Payable Aging Report."""
    context = _get_admin_base_context(_("AP Aging Report"), request)
    target_company: Optional[Company] = None
    today = date.today()

    # For form repopulation and defaults
    context['as_of_date_param'] = request.GET.get('as_of_date', today.isoformat())
    context['aging_bucket_days_config_str'] = request.GET.get(
        'aging_buckets',
        ",".join(map(str, reports_service.DEFAULT_AP_AGING_BUCKETS_DAYS))
    )
    context['report_data_available'] = False
    context['as_of_date_to_display'] = None # Initialize

    try:
        target_company = _get_company_for_report_or_raise(request)
        context['company'] = target_company
        context['title'] = f"{_('AP Aging Report')} - {target_company.name}"

        # Parse as_of_date from the parameter used for repopulation
        # This will be today's date on first load, or submitted date on subsequent loads.
        as_of_date_val = date.fromisoformat(context['as_of_date_param'])

        # Determine aging buckets to use
        buckets_str_from_get = request.GET.get('aging_buckets')
        if buckets_str_from_get:
            try:
                aging_buckets_to_use = [int(b.strip()) for b in buckets_str_from_get.split(',') if b.strip()]
                # Update config_str to reflect what was parsed from GET
                context['aging_bucket_days_config_str'] = buckets_str_from_get
            except ValueError:
                aging_buckets_to_use = reports_service.DEFAULT_AP_AGING_BUCKETS_DAYS
                messages.error(request, _("Invalid aging buckets format. Using default AP buckets."))
                # Reset config_str to default if parsing failed
                context['aging_bucket_days_config_str'] = ",".join(map(str, reports_service.DEFAULT_AP_AGING_BUCKETS_DAYS))
        else:
            # If not in GET, it implies initial load or user cleared it.
            # The value in context['aging_bucket_days_config_str'] is already set to default.
            aging_buckets_to_use = [int(b.strip()) for b in context['aging_bucket_days_config_str'].split(',') if b.strip()]


        # Generate report only if 'as_of_date' was explicitly in request.GET (form submission)
        # OR if 'company_id' was in request.GET (company changed, potentially keep old date)
        # AND target_company is successfully set.
        # The main trigger for generation is the 'View Report' button click which includes 'as_of_date'.
        if ('as_of_date' in request.GET or 'company_id' in request.GET) and target_company:
            context['as_of_date_to_display'] = as_of_date_val # Set for template display

            report_data = reports_service.generate_ap_aging_report(
                company_id=target_company.id,
                as_of_date=as_of_date_val,
                report_currency=getattr(target_company, 'default_currency_code', 'USD'),
                aging_buckets_days=aging_buckets_to_use
            )
            context.update(report_data) # Adds keys like 'aging_data', 'bucket_labels', etc.
            if report_data.get('aging_data'): # Check if there's actual data
                context['report_data_available'] = True
            # If no data, the template's "No outstanding payables..." message will show
        elif target_company and not ('as_of_date' in request.GET):
            # Company selected, but date not yet submitted for report generation.
            # The template condition `{% if company and not as_of_date_to_display %}` will handle the message.
            pass


    except ValueError as ve: # Catches date.fromisoformat errors or errors from _get_company_for_report_or_raise
        messages.error(request, str(ve))
        context['report_error'] = str(ve)
        context['as_of_date_to_display'] = None # Ensure this is not set on error
        if request.user.is_superuser and ("Company ID" in str(ve) or "select an active company" in str(ve).lower()):
            if Company: context['all_companies'] = Company.objects.filter(is_active=True).order_by('name')
            context['show_company_selector'] = True
    except (Http404, DjangoPermissionDenied) as e:
        messages.error(request, str(e))
        context['report_error'] = str(e)
    except ReportGenerationError as rge:
        messages.error(request, f"{_('Error generating AP Aging')}: {rge}")
        context['report_error'] = f"{_('Error generating AP Aging')}: {rge}"
    except Exception as e:
        logger.exception(f"Error generating admin AP Aging for Co '{getattr(target_company, 'name', 'N/A')}'")
        messages.error(request, _("Unexpected error generating AP Aging."))
        context['report_error'] = _("Unexpected error.")

    # Ensure 'opts' and other admin context vars are present
    if 'opts' not in context: context['opts'] = getattr(Company, '_meta', None) if Company else getattr(Account, '_meta', None)
    if request.user.is_superuser and 'all_companies' not in context and Company:
         context['all_companies'] = Company.objects.filter(is_active=True).order_by('name')
         context['show_company_selector'] = True


    return render(request, 'admin/crp_accounting/reports/ap_aging_report.html', context)


@staff_member_required
def admin_vendor_statement_view(request: HttpRequest,
                                supplier_pk: Optional[PK_TYPE] = None) -> HttpResponse:
    """Displays a Vendor/Supplier Statement."""
    context = _get_admin_base_context(_("Vendor Statement"), request)
    target_company: Optional[Company] = None
    target_supplier: Optional[Party] = None
    today = date.today()
    default_start_date = today.replace(day=1)
    next_month_start = (default_start_date.replace(day=28) + timedelta(days=4)).replace(day=1)
    default_end_date = next_month_start - timedelta(days=1)

    context['start_date_param'] = request.GET.get('start_date', default_start_date.isoformat())
    context['end_date_param'] = request.GET.get('end_date', default_end_date.isoformat())
    supplier_pk_from_get = request.GET.get('supplier_id')
    effective_supplier_pk = supplier_pk or supplier_pk_from_get
    context['selected_supplier_id_param'] = str(effective_supplier_pk) if effective_supplier_pk else ''

    context['report_data_available'] = False
    context['start_date_to_display'] = None
    context['end_date_to_display'] = None
    context['selected_supplier'] = None
    context['supplier'] = None
    context['suppliers_for_selection'] = [] # Initialize as empty list
    context['all_suppliers_for_company'] = [] # Initialize as empty list


    try:
        target_company = _get_company_for_report_or_raise(request)
        context['company'] = target_company
        context['title'] = f"{_('Vendor Statement')} - {target_company.name}"
        logger.info(f"[VendorStmtView] Target company: {target_company.name}")

        if target_company:
            suppliers_qs = Party.objects.filter(
                company=target_company,
                party_type=CorePartyType.SUPPLIER.value,
                is_active=True
            ).order_by('name')
            context['suppliers_for_selection'] = suppliers_qs
            context['all_suppliers_for_company'] = suppliers_qs
            logger.info(f"[VendorStmtView] Found {suppliers_qs.count()} suppliers for company {target_company.name}.")
            if not suppliers_qs.exists():
                logger.warning(f"[VendorStmtView] No active suppliers found for company {target_company.name}. Dropdown will be empty.")
                # messages.info(request, _("No active suppliers found for this company.")) # Optional message

        if effective_supplier_pk and context['start_date_param'] and context['end_date_param'] and target_company:
            logger.info(f"[VendorStmtView] Attempting to generate statement for supplier PK: {effective_supplier_pk}")
            target_supplier = get_object_or_404(
                Party.objects.select_related('company'),
                pk=effective_supplier_pk,
                company=target_company,
                party_type=CorePartyType.SUPPLIER.value
            )
            context['supplier'] = target_supplier
            context['selected_supplier'] = target_supplier
            context['title'] = f"{_('Statement')}: {target_supplier.name} - {target_company.name}"
            logger.info(f"[VendorStmtView] Target supplier: {target_supplier.name}")

            start_date_val = date.fromisoformat(context['start_date_param'])
            end_date_val = date.fromisoformat(context['end_date_param'])

            if start_date_val > end_date_val:
                raise DjangoValidationError(_("Start date cannot be after end date."))

            context['start_date_to_display'] = start_date_val
            context['end_date_to_display'] = end_date_val

            report_data = reports_service.generate_vendor_statement(
                company_id=target_company.id,
                supplier_id=target_supplier.pk,
                start_date=start_date_val,
                end_date=end_date_val,
                report_currency=getattr(target_company, 'default_currency_code', 'USD')
            )
            context.update(report_data)
            context['statement_data'] = report_data
            if report_data.get('lines') is not None:
                context['report_data_available'] = True
                logger.info(f"[VendorStmtView] Statement generated successfully for {target_supplier.name}.")
            else:
                logger.info(f"[VendorStmtView] Statement generation for {target_supplier.name} resulted in no lines.")


        elif target_company and not effective_supplier_pk and ('supplier_id' in request.GET or 'start_date' in request.GET or 'end_date' in request.GET):
            messages.info(request, _("Please select a supplier to view the statement."))
            logger.info("[VendorStmtView] Prompting user to select a supplier.")
        elif target_company and not effective_supplier_pk :
             # This case is when the page loads, company is known, but no supplier is selected yet.
             # The template should show "Please select a supplier..." if this is the case.
             logger.debug("[VendorStmtView] Company selected, no supplier chosen yet by user.")


    except DjangoValidationError as dve:
        error_msg = "; ".join(dve.messages) if hasattr(dve, 'messages') and isinstance(dve.messages, list) else str(dve)
        messages.error(request, error_msg)
        context['report_error'] = error_msg
        logger.warning(f"[VendorStmtView] Validation error: {error_msg}")
    except ValueError as ve: # Catches date.fromisoformat or errors from _get_company_for_report_or_raise
        messages.error(request, str(ve))
        context['report_error'] = str(ve)
        logger.warning(f"[VendorStmtView] Value error: {str(ve)}")
        if request.user.is_superuser and ("Company ID" in str(ve) or "select an active company" in str(ve).lower()):
            if Company: context['all_companies'] = Company.objects.filter(is_active=True).order_by('name')
            context['show_company_selector'] = True
        # If company is set but error occurred, ensure supplier dropdown is still populated
        elif target_company and not context.get('suppliers_for_selection'): # Check if already populated
            suppliers_qs_err = Party.objects.filter(company=target_company, party_type=CorePartyType.SUPPLIER.value, is_active=True).order_by('name')
            context['suppliers_for_selection'] = suppliers_qs_err
            context['all_suppliers_for_company'] = suppliers_qs_err
            logger.info(f"[VendorStmtView] Repopulated suppliers for {target_company.name} after ValueError.")
    except Http404:
        messages.error(request, _("Supplier or Company not found. Please check your selection or permissions."))
        context['report_error'] = _("Supplier or Company not found.")
        logger.warning("[VendorStmtView] Http404 encountered (Supplier or Company not found).")
        if target_company and not context.get('suppliers_for_selection'):
            suppliers_qs_404 = Party.objects.filter(company=target_company, party_type=CorePartyType.SUPPLIER.value, is_active=True).order_by('name')
            context['suppliers_for_selection'] = suppliers_qs_404
            context['all_suppliers_for_company'] = suppliers_qs_404
            logger.info(f"[VendorStmtView] Repopulated suppliers for {target_company.name} after Http404.")
    except DjangoPermissionDenied as e:
        messages.error(request, str(e))
        context['report_error'] = str(e)
        logger.warning(f"[VendorStmtView] Permission denied: {str(e)}")
    except ReportGenerationError as rge:
        messages.error(request, f"{_('Error generating statement')}: {rge}")
        context['report_error'] = f"{_('Error generating statement')}: {rge}"
        logger.error(f"[VendorStmtView] ReportGenerationError: {rge}")
    except Exception as e:
        logger.exception(
            f"[VendorStmtView] Unexpected error for Co '{getattr(target_company, 'name', 'N/A')}', Supp '{getattr(target_supplier, 'name', effective_supplier_pk)}'")
        messages.error(request, _("Unexpected error generating statement."))
        context['report_error'] = _("Unexpected error.")

    if 'opts' not in context: context['opts'] = getattr(Party, '_meta', None) # Party model meta
    if request.user.is_superuser and 'all_companies' not in context and Company:
         context['all_companies'] = Company.objects.filter(is_active=True).order_by('name')
         context['show_company_selector'] = True

    return render(request, 'admin/crp_accounting/reports/vendor_statement_report.html', context)
# =============================================================================
# NEW: Accounts Payable (AP) Download Views (Tenant Aware)
# =============================================================================

@staff_member_required
def download_ap_aging_excel(request: HttpRequest) -> HttpResponse:
    target_company: Optional[Company] = None
    if not OPENPYXL_AVAILABLE: return HttpResponse(str(_("Excel export library is missing.")), status=501)
    try:
        target_company = _get_company_for_report_or_raise(request)
        date_params = _get_validated_date_params_for_download(request, ['as_of_date'])
        as_of_date_val = date_params['as_of_date']

        buckets_str = request.GET.get('aging_buckets')
        aging_buckets_to_use = [int(b.strip()) for b in buckets_str.split(
            ',')] if buckets_str else reports_service.DEFAULT_AR_AGING_BUCKETS_DAYS  # Reusing AR default for now

        report_data = reports_service.generate_ap_aging_report(
            company_id=target_company.id, as_of_date=as_of_date_val,
            report_currency=getattr(target_company, 'default_currency_code', 'USD'),
            aging_buckets_days=aging_buckets_to_use
        )

        workbook = openpyxl.Workbook();
        sheet = workbook.active;
        sheet.title = str(_("AP Aging"))

        # CORRECTED currency_symbol logic
        company_sym = getattr(target_company, 'default_currency_symbol', None)
        report_sym = report_data.get('report_currency', None)
        if company_sym is not None:
            currency_symbol = company_sym
        elif report_sym is not None:
            currency_symbol = report_sym
        else:
            currency_symbol = ''

        number_format_currency = f'"{currency_symbol}"#,##0.00;[Red]-"{currency_symbol}"#,##0.00'

        # Example Header
        sheet['A1'] = f"{str(_('AP Aging Report'))} - {target_company.name}"
        sheet.merge_cells(start_row=1, start_column=1, end_row=1,
                          end_column=len(report_data.get('bucket_labels', [])) + 2)
        sheet['A1'].font = Font(bold=True, size=14);
        sheet['A1'].alignment = Alignment(horizontal='center')
        sheet['A2'] = f"{str(_('As of Date:'))} {as_of_date_val.strftime('%B %d, %Y')}"
        sheet.merge_cells(start_row=2, start_column=1, end_row=2,
                          end_column=len(report_data.get('bucket_labels', [])) + 2)
        sheet['A2'].alignment = Alignment(horizontal='center');
        sheet.append([])

        excel_headers = [str(_("Supplier Name"))] + [str(label) for label in report_data.get('bucket_labels', [])] + [
            str(_("Total Due"))]
        sheet.append(excel_headers)
        header_row_num = sheet.max_row
        for col_idx, header_title in enumerate(excel_headers, 1):
            cell = sheet.cell(row=header_row_num, column=col_idx, value=header_title)
            cell.font = Font(bold=True)
            cell.border = Border(bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal='left' if col_idx == 1 else 'right')
            sheet.column_dimensions[get_column_letter(col_idx)].width = 30 if col_idx == 1 else 18

        for entry in report_data.get('aging_data', []):  # aging_data from AP report will have 'supplier_name'
            row_values = [str(entry.get('party_name', 'N/A'))]  # Use 'party_name' from AgingEntry
            for bucket_label in report_data.get('bucket_labels', []): row_values.append(
                entry.get('buckets', {}).get(bucket_label, ZERO))
            row_values.append(entry.get('total_due', ZERO))
            sheet.append(row_values)
            current_data_row = sheet.max_row
            for col_idx in range(2, len(excel_headers) + 1):  # Start from 2nd col for amounts
                sheet.cell(row=current_data_row, column=col_idx).number_format = number_format_currency
                sheet.cell(row=current_data_row, column=col_idx).alignment = Alignment(horizontal='right')

        sheet.append([])  # Empty row before totals

        grand_total_row_values = [str(_("Grand Totals"))]
        for bucket_label in report_data.get('bucket_labels', []):
            grand_total_row_values.append(report_data.get('grand_totals_by_bucket', {}).get(bucket_label, ZERO))
        grand_total_row_values.append(
            report_data.get('grand_total_due_all_customers', ZERO))  # Field name might be different for AP
        sheet.append(grand_total_row_values)
        total_row_num = sheet.max_row
        for col_idx in range(1, len(grand_total_row_values) + 1):
            cell = sheet.cell(row=total_row_num, column=col_idx)
            cell.font = Font(bold=True)
            if col_idx > 1:
                cell.number_format = number_format_currency
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"{target_company.name}_AP_Aging_{as_of_date_val.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"';
        workbook.save(response);
        return response
    except ValueError as ve:
        return HttpResponseBadRequest(str(ve))
    except (Http404, DjangoPermissionDenied) as e:
        return HttpResponse(str(e), status=403 if isinstance(e, DjangoPermissionDenied) else 404)
    except ReportGenerationError as rge:
        return HttpResponse(f"{str(_('Report error'))}: {rge}", status=500)
    except Exception as e:
        logger.exception(
            f"Error exporting AP Aging Excel for Co '{getattr(target_company, 'name', 'N/A')}'");
        return HttpResponse(
            str(_("Excel generation error.")), status=500)



@staff_member_required
def download_vendor_statement_excel(request: HttpRequest, supplier_pk: PK_TYPE) -> HttpResponse:
    """
    Generates and serves a Vendor Statement as an Excel file.
    """
    target_company: Optional[Company] = None
    target_supplier: Optional[Party] = None
    if not OPENPYXL_AVAILABLE:
        return HttpResponse(str(_("Excel export library (openpyxl) is not installed.")), status=501)

    try:
        # 1. Get company, supplier, and date parameters (similar to the PDF function)
        target_company = _get_company_for_report_or_raise(request)
        date_params = _get_validated_date_params_for_download(request, ['start_date', 'end_date'])
        start_date_val, end_date_val = date_params['start_date'], date_params['end_date']
        if start_date_val > end_date_val:
            raise ValueError(str(_("Start date cannot be after end date.")))

        target_supplier = get_object_or_404(
            Party.objects.select_related('company'),
            pk=supplier_pk,
            company=target_company,
            party_type=CorePartyType.SUPPLIER.value
        )

        # 2. Generate the statement data using your existing service
        report_data = reports_service.generate_vendor_statement(
            company_id=target_company.id,
            supplier_id=target_supplier.pk,
            start_date=start_date_val,
            end_date=end_date_val,
            report_currency=getattr(target_company, 'default_currency_code', 'USD')
        )

        # 3. Create the Excel workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = str(_("Vendor Statement"))

        # Setup currency formatting
        currency_symbol = report_data.get('report_currency_symbol', '$')
        number_format_currency = f'"{currency_symbol}"#,##0.00;[Red]-"{currency_symbol}"#,##0.00'

        # 4. Write Headers and Titles
        sheet['A1'] = f"{str(_('Statement for Supplier:'))} {target_supplier.name}"
        sheet.merge_cells('A1:F1')
        sheet['A1'].font = Font(bold=True, size=16)

        sheet['A2'] = f"{str(_('Company:'))} {target_company.name}"
        sheet.merge_cells('A2:F2')

        sheet[
            'A3'] = f"{str(_('Period:'))} {start_date_val.strftime('%d-%m-%Y')} {str(_('to'))} {end_date_val.strftime('%d-%m-%Y')}"
        sheet.merge_cells('A3:F3')

        sheet.append([])  # Blank row

        # Table Headers
        headers = [
            str(_("Date")), str(_("Transaction Type")), str(_("Reference")),
            str(_("Payment / Debit Note")), str(_("Bill / Credit Note")), str(_("Balance Due to Supplier"))
        ]
        sheet.append(headers)
        header_row_num = sheet.max_row
        for col_idx, cell_value in enumerate(headers, 1):
            cell = sheet.cell(row=header_row_num, column=col_idx)
            cell.font = Font(bold=True)
            cell.border = Border(bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal='right' if col_idx > 3 else 'left')
            # Set column widths
            if col_idx == 1: sheet.column_dimensions[get_column_letter(col_idx)].width = 15
            if col_idx == 2: sheet.column_dimensions[get_column_letter(col_idx)].width = 25
            if col_idx == 3: sheet.column_dimensions[get_column_letter(col_idx)].width = 20
            if col_idx in [4, 5, 6]: sheet.column_dimensions[get_column_letter(col_idx)].width = 22

        # 5. Write Data Rows
        # Opening Balance
        opening_balance_row = [
            start_date_val.strftime('%d-%m-%Y'),
            str(_("Opening Balance")),
            None, None, None,
            report_data.get('opening_balance', ZERO)
        ]
        sheet.append(opening_balance_row)
        sheet.cell(row=sheet.max_row, column=6).number_format = number_format_currency

        # Transaction Lines
        for line in report_data.get('lines', []):
            row_data = [
                line.get('date'),
                line.get('transaction_type'),
                line.get('reference'),
                line.get('payment_or_debit'),
                line.get('bill_or_credit'),
                line.get('balance')
            ]
            sheet.append(row_data)
            current_row = sheet.max_row
            # Apply currency format to amount columns
            sheet.cell(row=current_row, column=4).number_format = number_format_currency
            sheet.cell(row=current_row, column=5).number_format = number_format_currency
            sheet.cell(row=current_row, column=6).number_format = number_format_currency

        # Closing Balance
        closing_balance_row = [
            None, None, None, None,
            str(_("Closing Balance as of")) + f" {end_date_val.strftime('%d-%m-%Y')}",
            report_data.get('closing_balance', ZERO)
        ]
        sheet.append(closing_balance_row)
        closing_row_num = sheet.max_row
        closing_label_cell = sheet.cell(row=closing_row_num, column=5)
        closing_amount_cell = sheet.cell(row=closing_row_num, column=6)
        closing_label_cell.font = Font(bold=True)
        closing_amount_cell.font = Font(bold=True)
        closing_amount_cell.number_format = number_format_currency
        closing_amount_cell.alignment = Alignment(horizontal='right')

        # 6. Create the HTTP response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Vendor_Statement_{target_supplier.name.replace(' ', '_')}_{end_date_val.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response

    except ValueError as ve:
        return HttpResponseBadRequest(str(ve))
    except (Http404, DjangoPermissionDenied) as e:
        return HttpResponse(str(e), status=403 if isinstance(e, DjangoPermissionDenied) else 404)
    except ReportGenerationError as rge:
        return HttpResponse(f"{str(_('Report generation error'))}: {rge}", status=500)
    except Exception as e:
        logger.exception(
            f"Error exporting Vendor Statement Excel for Co '{getattr(target_company, 'name', 'N/A')}', Supp '{getattr(target_supplier, 'name', supplier_pk)}'")
        return HttpResponse(str(_("An unexpected error occurred during Excel generation.")), status=500)


@staff_member_required
def download_trial_balance_pdf(request: HttpRequest) -> HttpResponse:
    """
    Generates and serves the Trial Balance report as a clean PDF file.
    """
    if not XHTML2PDF_AVAILABLE:
        return HttpResponse(str(_("PDF export library is not installed.")), status=501)
    try:
        target_company = _get_company_for_report_or_raise(request)
        date_params = _get_validated_date_params_for_download(request, ['as_of_date'])
        as_of_date_val = date_params['as_of_date']

        report_data = reports_service.generate_trial_balance_structured(
            company_id=target_company.id, as_of_date=as_of_date_val
        )
        context = {
            'company': target_company,
            'report_title': _("Trial Balance Report"),
            'as_of_date_param': as_of_date_val,  # **FIX:** Pass the date object directly
            **report_data
        }

        # **FIX:** Point to the dedicated PDF template
        pdf_template_path = 'admin/crp_accounting/reports/trial_balance_pdf.html'
        pdf_buffer = _render_to_pdf(pdf_template_path, context, request=request)

        response = HttpResponse(pdf_buffer, content_type='application/pdf')
        filename = f"{target_company.name}_Trial_Balance_{as_of_date_val.strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response
    except Exception as e:
        logger.exception("Error generating Trial Balance PDF")
        return HttpResponse(f"Error: {e}", status=500)


@staff_member_required
def download_profit_loss_pdf(request: HttpRequest) -> HttpResponse:
    """
    Generates and serves the Profit & Loss Statement as a clean PDF file.
    """
    if not XHTML2PDF_AVAILABLE:
        return HttpResponse(str(_("PDF export library is not installed.")), status=501)
    try:
        target_company = _get_company_for_report_or_raise(request)
        date_params = _get_validated_date_params_for_download(request, ['start_date', 'end_date'])
        start_date_val, end_date_val = date_params['start_date'], date_params['end_date']

        report_data = reports_service.generate_profit_loss(
            company_id=target_company.id, start_date=start_date_val, end_date=end_date_val
        )
        context = {
            'company': target_company,
            'report_title': _("Profit & Loss Statement"),
            'start_date_param': start_date_val,  # **FIX:** Pass date object
            'end_date_param': end_date_val,  # **FIX:** Pass date object
            **report_data
        }

        # **FIX:** Point to the dedicated PDF template
        pdf_template_path = 'admin/crp_accounting/reports/profit_loss_pdf.html'
        pdf_buffer = _render_to_pdf(pdf_template_path, context, request=request)

        response = HttpResponse(pdf_buffer, content_type='application/pdf')
        filename = f"{target_company.name}_Profit_Loss_{end_date_val.strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response
    except Exception as e:
        logger.exception("Error generating P&L PDF")
        return HttpResponse(f"Error: {e}", status=500)


@staff_member_required
def download_balance_sheet_pdf(request: HttpRequest) -> HttpResponse:
    """
    Generates and serves the Balance Sheet as a clean PDF file using a
    professional side-by-side layout.
    """
    if not XHTML2PDF_AVAILABLE:
        return HttpResponse(str(_("PDF export library (xhtml2pdf) is not installed.")), status=501)

    target_company: Optional[Company] = None
    try:
        # 1. Get Company and Date parameters
        target_company = _get_company_for_report_or_raise(request)
        date_params = _get_validated_date_params_for_download(request, ['as_of_date'])
        as_of_date_val = date_params['as_of_date']

        # 2. Generate the core report data from your service
        report_data = reports_service.generate_balance_sheet(
            company_id=target_company.id,
            as_of_date=as_of_date_val,
            report_currency=getattr(target_company, 'default_currency_code', 'USD')
        )

        # ====================================================================
        # === THE FIX IS HERE: Calculate and add missing totals to the data ===
        # ====================================================================
        # Your service provides the main sections, but we need to compute the
        # final combined total here before sending it to the template.
        total_liabilities = report_data.get('liabilities', {}).get('total', ZERO)
        total_equity = report_data.get('equity', {}).get('total', ZERO)

        # Add the new key to the report_data dictionary
        report_data['total_liabilities_and_equity'] = total_liabilities + total_equity
        # Also ensure the currency symbol is available for the template
        if 'report_currency_symbol' not in report_data:
            report_data['report_currency_symbol'] = getattr(target_company, 'default_currency_symbol', '$')
        # ====================================================================

        # 3. Prepare the full context for the template
        context = {
            'company': target_company,
            'report_title': _("Balance Sheet"),
            'as_of_date_param': as_of_date_val,  # Pass the actual date object
            **report_data  # Unpack all keys from report_data into the context
        }

        # 4. Point to the dedicated PDF template and render it
        pdf_template_path = 'admin/crp_accounting/reports/balance_sheet_pdf.html'
        pdf_buffer = _render_to_pdf(pdf_template_path, context, request=request)

        # 5. Create and return the final HTTP response
        response = HttpResponse(pdf_buffer, content_type='application/pdf')
        filename = f"{target_company.name}_Balance_Sheet_{as_of_date_val.strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response

    except (ValueError, Http404, DjangoPermissionDenied) as e:
        # Handle expected errors gracefully
        return HttpResponse(str(e), status=400)
    except ReportGenerationError as rge:
        logger.error(f"Report generation error for Balance Sheet PDF: {rge}")
        return HttpResponse(f"Error generating report: {rge}", status=500)
    except Exception as e:
        # Log unexpected errors
        logger.exception(
            f"Unexpected error generating Balance Sheet PDF for Co '{getattr(target_company, 'name', 'N/A')}'")
        return HttpResponse(str(_("An unexpected PDF generation error occurred.")), status=500)


@staff_member_required
def download_ar_aging_pdf(request: HttpRequest) -> HttpResponse:
    """
    Generates and serves the AR Aging report as a clean PDF file.
    (Definitive version, written to match the service's output)
    """
    if not XHTML2PDF_AVAILABLE:
        return HttpResponse(str(_("PDF export library is not installed.")), status=501)

    try:
        # 1. Get parameters
        target_company = _get_company_for_report_or_raise(request)
        date_params = _get_validated_date_params_for_download(request, ['as_of_date'])
        as_of_date_val = date_params['as_of_date']

        buckets_str = request.GET.get('aging_buckets')
        aging_buckets_to_use = [int(b.strip()) for b in
                                buckets_str.split(',')] if buckets_str else None  # Pass None to use default

        # 2. Get data from your service
        report_data = reports_service.generate_ar_aging_report(
            company_id=target_company.id,
            as_of_date=as_of_date_val,
            aging_buckets_days=aging_buckets_to_use,
            report_currency=getattr(target_company, 'default_currency_code', 'USD')
        )

        # 3. Build the context for the template
        # This context now directly uses the keys from your service's return dictionary.
        context = {
            'company': target_company,
            'report_title': _("AR Aging Report"),
            'as_of_date_param': as_of_date_val,
            'report_currency_symbol': report_data.get('report_currency_symbol', '$'),
            'bucket_labels': report_data.get('bucket_labels', []),
            'aging_data': report_data.get('aging_data', []),
            'grand_totals_by_bucket': report_data.get('grand_totals_by_bucket', {}),
            'grand_total_due_all_customers': report_data.get('grand_total_due_all_customers', 0),
        }

        # 4. Render the PDF using the dedicated template
        pdf_template_path = 'admin/crp_accounting/reports/ar_aging_pdf.html'
        pdf_buffer = _render_to_pdf(pdf_template_path, context, request=request)

        response = HttpResponse(pdf_buffer, content_type='application/pdf')
        filename = f"{target_company.name}_AR_Aging_{as_of_date_val.strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response

    except (ValueError, Http404, DjangoPermissionDenied) as e:
        return HttpResponse(str(e), status=400)
    except Exception as e:
        logger.exception("Error generating AR Aging PDF")
        return HttpResponse(f"Error: {e}", status=500)


@staff_member_required
def download_customer_statement_pdf(request: HttpRequest, customer_pk: PK_TYPE) -> HttpResponse:
    """
    Generates and serves the Customer Statement as a clean PDF file. (Corrected Currency)
    """
    if not XHTML2PDF_AVAILABLE:
        return HttpResponse(str(_("PDF export library is not installed.")), status=501)

    try:
        # ... (Get parameters and report_data - this part is correct) ...
        target_company = _get_company_for_report_or_raise(request)
        target_customer = get_object_or_404(Party, pk=customer_pk, company=target_company,
                                            party_type=CorePartyType.CUSTOMER.value)
        date_params = _get_validated_date_params_for_download(request, ['start_date', 'end_date'])
        start_date_val, end_date_val = date_params['start_date'], date_params['end_date']

        report_data = reports_service.generate_customer_statement(
            company_id=target_company.id,
            customer_id=target_customer.pk,
            start_date=start_date_val,
            end_date=end_date_val,
            report_currency=getattr(target_company, 'default_currency_code', 'USD')
        )

        # =======================================================================
        # === THE CURRENCY FIX IS HERE                                        ===
        # =======================================================================
        # Determine the correct currency symbol to use, defaulting to the code.
        report_currency = report_data.get('report_currency', 'INR')
        currency_symbol = report_data.get('report_currency_symbol')
        if not currency_symbol:
            # If no symbol is provided (like for INR), we will just show the code.
            # Or, leave it blank if you prefer no prefix at all.
            currency_symbol = ""  # Let's use no symbol to match your web view.
        # =======================================================================

        context = {
            'company': target_company,
            'customer': target_customer,
            'report_title': _("Customer Statement"),
            'start_date_param': start_date_val,
            'end_date_param': end_date_val,
            'lines': report_data.get('lines', []),
            'opening_balance': report_data.get('opening_balance', 0),
            'closing_balance': report_data.get('closing_balance', 0),
            'report_currency': report_currency,
            'report_currency_symbol': currency_symbol,  # Pass the corrected symbol
        }

        pdf_template_path = 'admin/crp_accounting/reports/customer_statement_pdf.html'
        pdf_buffer = _render_to_pdf(pdf_template_path, context, request=request)

        response = HttpResponse(pdf_buffer, content_type='application/pdf')
        filename = f"Statement_{target_customer.name.replace(' ', '_')}.pdf"
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response

    except Exception as e:
        logger.exception("Error generating Customer Statement PDF")
        return HttpResponse(f"Error: {e}", status=500)
@staff_member_required
def download_ap_aging_pdf(request: HttpRequest) -> HttpResponse:
    """
    Generates and serves the AP Aging report as a clean PDF file.
    """
    if not XHTML2PDF_AVAILABLE:
        return HttpResponse(str(_("PDF export library is not installed.")), status=501)
    try:
        target_company = _get_company_for_report_or_raise(request)
        date_params = _get_validated_date_params_for_download(request, ['as_of_date'])
        as_of_date_val = date_params['as_of_date']

        buckets_str = request.GET.get('aging_buckets')
        aging_buckets_to_use = [int(b.strip()) for b in buckets_str.split(
            ',')] if buckets_str else reports_service.DEFAULT_AP_AGING_BUCKETS_DAYS

        report_data = reports_service.generate_ap_aging_report(
            company_id=target_company.id, as_of_date=as_of_date_val, aging_buckets_days=aging_buckets_to_use
        )
        context = {
            'company': target_company,
            'report_title': _("AP Aging Report"),
            'as_of_date_param': as_of_date_val,  # **FIX:** Pass date object
            **report_data
        }

        # **FIX:** Point to the dedicated PDF template
        pdf_template_path = 'admin/crp_accounting/reports/ap_aging_pdf.html'
        pdf_buffer = _render_to_pdf(pdf_template_path, context, request=request)

        response = HttpResponse(pdf_buffer, content_type='application/pdf')
        filename = f"{target_company.name}_AP_Aging_{as_of_date_val.strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response
    except Exception as e:
        logger.exception("Error generating AP Aging PDF")
        return HttpResponse(f"Error: {e}", status=500)


@staff_member_required
def download_vendor_statement_pdf(request: HttpRequest, supplier_pk: PK_TYPE) -> HttpResponse:
    """
    Generates and serves the Vendor Statement as a clean PDF file.
    """
    if not XHTML2PDF_AVAILABLE:
        return HttpResponse(str(_("PDF export library is not installed.")), status=501)
    try:
        target_company = _get_company_for_report_or_raise(request)
        target_supplier = get_object_or_404(Party, pk=supplier_pk, company=target_company,
                                            party_type=CorePartyType.SUPPLIER.value)
        date_params = _get_validated_date_params_for_download(request, ['start_date', 'end_date'])
        start_date_val, end_date_val = date_params['start_date'], date_params['end_date']

        report_data = reports_service.generate_vendor_statement(
            company_id=target_company.id, supplier_id=target_supplier.pk, start_date=start_date_val,
            end_date=end_date_val
        )
        context = {
            'company': target_company,
            'supplier': target_supplier,
            'report_title': _("Vendor Statement"),
            'start_date_param': start_date_val,  # **FIX:** Pass date object
            'end_date_param': end_date_val,  # **FIX:** Pass date object
            **report_data
        }

        # **FIX:** Point to the dedicated PDF template
        pdf_template_path = 'admin/crp_accounting/reports/vendor_statement_pdf.html'
        pdf_buffer = _render_to_pdf(pdf_template_path, context, request=request)

        response = HttpResponse(pdf_buffer, content_type='application/pdf')
        filename = f"Vendor_Statement_{target_supplier.name.replace(' ', '_')}.pdf"
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response
    except Exception as e:
        logger.exception("Error generating Vendor Statement PDF")
        return HttpResponse(f"Error: {e}", status=500)

@staff_member_required
def admin_reports_hub_view(request: HttpRequest) -> HttpResponse:
    context = _get_admin_base_context(_("Accounting Reports Hub"), request)
    user_name_attr = getattr(request.user, 'name', None)  # Try 'name' attribute first
    user_name = user_name_attr if user_name_attr else getattr(request.user, 'username', str(request.user.id))

    logger.debug(f"[ReportsHubView] User: {user_name}, Is Superuser: {request.user.is_superuser}")
    if request.user.is_superuser:
        if Company:
            companies_qs = Company.objects.filter(is_active=True).order_by('name')
            context['all_companies'] = companies_qs
            logger.debug(f"[ReportsHubView] Found companies for SU: {[c.name for c in companies_qs]}")
        else:
            logger.warning("[ReportsHubView] Company model not available, cannot list companies for SU.")
            context['all_companies'] = []
    else:
        logger.debug(f"[ReportsHubView] User {user_name} is not SU, not populating all_companies.")
    return render(request, 'admin/crp_accounting/reports_hub.html', context)